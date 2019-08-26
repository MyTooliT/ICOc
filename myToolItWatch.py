import sys
import os

import xml.etree.ElementTree as ET
import PeakCanFd
from PeakCanFd import rreplace 
from MyToolItNetworkNumbers import *
from MyToolItCommands import *
from time import sleep, time
from random import randint
from MyToolItSth import *
from datetime import datetime
import getopt
import openpyxl
from openpyxl.styles import Font
import copy
import argparse
import multiprocessing
from Plotter import vPlotter, tArray2Binary
import socket


HOST = 'localhost'  # The remote host
PORT = 50007  # The same port as used by the server

BlueToothDeviceListAquireTime = 5


def to8bitSigned(num): 
    mask7 = 128  # Check 8th bit ~ 2^8
    mask2s = 127  # Keep first 7 bits
    if (mask7 & num == 128):  # Check Sign (8th bit)
        num = -((~int(num) + 1) & mask2s)  # 2's complement
    return num


def messageValueGet(m):        
    Acc = ((0xFF & m[1]) << 8) | (0xFF & m[0])
    return Acc  


Watch = {
    "IntervalDimMinX" : 10,  # Minimum interval time in ms
    "DisplayTimeMax" : 10,  # Maximum Display Time in seconds
    "DisplaySampleRateMs" : 100,  # Maximum Display Time in seconds
    "DisplayBlockSize" : 10,
}

    
class myToolItWatch():

    def __init__(self):
        self.KeyBoadInterrupt = False  
        self.bError = False     
        self.bClose = True   
        self.bConnected = False        
        self.iMsgLoss = 0
        self.iMsgsTotal = 0
        self.iMsgCounterLast = 0
        self.PeakCan = PeakCanFd.PeakCanFd(PeakCanFd.PCAN_BAUD_1M, "init.txt", "initError.txt", MyToolItNetworkNr["SPU1"], MyToolItNetworkNr["STH1"])
        self.vSave2Xml(False)
        self.vSthAutoConnect(False)
        self.PeakCan.Logger.Info("Start Time: " + self.sDateClock())
        self.bLogSet("init.txt")
        self.vConfigFileSet('configKeys.xml')
        self.bSampleSetupSet(None)
        self.vConfigSet(None, None)
        self.vSheetFileSet(None)
        self.vAccSet(True, False, False, 3)
        self.vVoltageSet(False, False, False, 3)
        self.vDeviceNameSet(TestConfig["DevName"])
        self.vDeviceAddressSet("0")
        self.vAdcConfig(2, 8, 64)
        self.vAdcRefVConfig("VDD")
        self.vDisplayTime(10)  
        self.vRunTime(10, 0)
        self.vGraphInit(Watch["DisplaySampleRateMs"], Watch["DisplayBlockSize"])
        self.PeakCan.readThreadStop()
            
    def __exit__(self):
        self.guiProcessStop()
        self.PeakCan.ReadThreadReset()
        if False != self.PeakCan.bConnected:
            self._BlueToothStatistics()
            ReceiveFailCounter = self._RoutingInformation()
            self._statusWords()
            self.PeakCan.BlueToothDisconnect(MyToolItNetworkNr["STU1"])
            if(0 < ReceiveFailCounter):
                self.bError = True
            self.PeakCan.readThreadStop()
        self.PeakCan.Logger.Info("End Time Stamp")
        
        if(False != self.bError):
            self.PeakCan.Logger.Info("bError")
            print("bError")
        else:
            self.PeakCan.Logger.Info("Fin")
        self.PeakCan.__exit__()  
        if(False != self.bError):
            raise
        if False != self.bSave:
            self.xmlSave()
        print("Fin")

    def _statusWords(self):
        ErrorWord = SthErrorWord()
        psw0 = self.PeakCan.statusWord0(MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("STH Status Word: " + hex(psw0))
        psw0 = self.PeakCan.statusWord0(MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STU Status Word: " + hex(psw0))
        ErrorWord.asword = self.PeakCan.statusWord1(MyToolItNetworkNr["STH1"])
        if True == ErrorWord.b.bAdcOverRun:
            self.bError = True
        self.PeakCan.Logger.Info("STH bError Word: " + hex(ErrorWord.asword))
        ErrorWord.asword = self.PeakCan.statusWord1(MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STU bError Word: " + hex(ErrorWord.asword))
        
    def _BlueToothStatistics(self):
        SendCounter = self.PeakCan.BlueToothCmd(MyToolItNetworkNr["STH1"], SystemCommandBlueTooth["SendCounter"])
        self.PeakCan.Logger.Info("BlueTooth Send Counter(STH1): " + str(SendCounter))
        Rssi = self.PeakCan.BlueToothRssi(MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("BlueTooth Rssi(STH1): " + str(Rssi) + "dBm")
        SendCounter = self.PeakCan.BlueToothCmd(MyToolItNetworkNr["STU1"], SystemCommandBlueTooth["SendCounter"])
        self.PeakCan.Logger.Info("BlueTooth Send Counter(STU1): " + str(SendCounter))
        ReceiveCounter = self.PeakCan.BlueToothCmd(MyToolItNetworkNr["STU1"], SystemCommandBlueTooth["ReceiveCounter"])
        self.PeakCan.Logger.Info("BlueTooth Receive Counter(STU1): " + str(ReceiveCounter))
        Rssi = self.PeakCan.BlueToothRssi(MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("BlueTooth Rssi(STU1): " + str(Rssi) + "dBm")

    def _RoutingInformationSthSend(self):
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STH1"], SystemCommandRouting["SendCounter"], MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STH1 - Send Counter(Port STU1): " + str(SendCounter))
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STH1"], SystemCommandRouting["SendFailCounter"], MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STH1 - Send Fail Counter(Port STU1): " + str(SendCounter))
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STH1"], SystemCommandRouting["SendLowLevelByteCounter"], MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STH1 - Send Byte Counter(Port STU1): " + str(SendCounter))

    def _RoutingInformationSthReceive(self):
        ReceiveCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STH1"], SystemCommandRouting["ReceiveCounter"], MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STH1 - Receive Counter(Port STU1): " + str(ReceiveCounter))
        ReceiveFailCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STH1"], SystemCommandRouting["ReceiveFailCounter"], MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STH1 - Receive Fail Counter(Port STU1): " + str(ReceiveFailCounter))
        ReceiveCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STH1"], SystemCommandRouting["ReceiveLowLevelByteCounter"], MyToolItNetworkNr["STU1"])
        self.PeakCan.Logger.Info("STH1 - Receive Byte Counter(Port STU1): " + str(ReceiveCounter))
        return ReceiveFailCounter

    def _RoutingInformationSth(self):
        self._RoutingInformationSthSend()
        ReceiveFailCounter = self._RoutingInformationSthReceive()
        return ReceiveFailCounter

    def _RoutingInformationStuPortSpuSend(self):
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["SendCounter"], MyToolItNetworkNr["SPU1"])
        self.PeakCan.Logger.Info("STU1 - Send Counter(Port SPU1): " + str(SendCounter))
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["SendFailCounter"], MyToolItNetworkNr["SPU1"])
        self.PeakCan.Logger.Info("STU1 - Send Fail Counter(Port SPU1): " + str(SendCounter))
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["SendLowLevelByteCounter"], MyToolItNetworkNr["SPU1"])
        self.PeakCan.Logger.Info("STU1 - Send Byte Counter(Port SPU1): " + str(SendCounter))

    def _RoutingInformationStuPortSpuReceive(self):
        ReceiveCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveCounter"], MyToolItNetworkNr["SPU1"])
        self.PeakCan.Logger.Info("STU1 - Receive Counter(Port SPU1): " + str(ReceiveCounter))
        ReceiveFailCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveFailCounter"], MyToolItNetworkNr["SPU1"])
        self.PeakCan.Logger.Info("STU1 - Receive Fail Counter(Port SPU1): " + str(ReceiveFailCounter))
        ReceiveCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveLowLevelByteCounter"], MyToolItNetworkNr["SPU1"])
        self.PeakCan.Logger.Info("STU1 - Receive Byte Counter(Port SPU1): " + str(ReceiveCounter))
        return ReceiveFailCounter

    def _RoutingInformationStuPortSpu(self):
        self._RoutingInformationStuPortSpuSend()
        ReceiveFailCounter = self._RoutingInformationStuPortSpuReceive()
        return ReceiveFailCounter

    def _RoutingInformationStuPortSthSend(self):
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["SendCounter"], MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("STU1 - Send Counter(Port STH1): " + str(SendCounter))
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["SendFailCounter"], MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("STU1 - Send Fail Counter(Port STH1): " + str(SendCounter))
        SendCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["SendLowLevelByteCounter"], MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("STU1 - Send Byte Counter(Port STH1): " + str(SendCounter))

    def _RoutingInformationStuPortSthReceive(self):
        ReceiveCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveCounter"], MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("STU1 - Receive Counter(Port STH1): " + str(ReceiveCounter))
        ReceiveFailCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveFailCounter"], MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("STU1 - Receive Fail Counter(Port STH1): " + str(ReceiveFailCounter))
        ReceiveCounter = self.PeakCan.RoutingInformationCmd(MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveLowLevelByteCounter"], MyToolItNetworkNr["STH1"])
        self.PeakCan.Logger.Info("STU1 - Receive Byte Counter(Port STH1): " + str(ReceiveCounter))
        return ReceiveFailCounter

    def _RoutingInformationStuPortSth(self):
        self._RoutingInformationStuPortSthSend()
        ReceiveFailCounter = self._RoutingInformationStuPortSthReceive()
        return ReceiveFailCounter

    def _RoutingInformation(self):
        ReceiveFailCounter = self._RoutingInformationSth()
        ReceiveFailCounter += self._RoutingInformationStuPortSth()
        ReceiveFailCounter += self._RoutingInformationStuPortSpu()
        return ReceiveFailCounter

# Setter Methods

    def vSave2Xml(self, bSave):
        self.bSave = bSave
        
    def vConfigFileSet(self, sfileName):
        self.sXmlFileName = sfileName
        self.tree = ET.parse(self.sXmlFileName)
        self.root = self.tree.getroot()
        
    def vConfigSet(self, product, sConfig):
        self.sProduct = None
        self.sConfig = None
        if "STH" == product:
            self.sProduct = "STH"
            self.PeakCan.vSetReceiver(MyToolItNetworkNr["STH1"])    
            self.sConfig = sConfig  
        elif "STU" == product: 
            self.sProduct = "STU"
            self.PeakCan.vSetReceiver(MyToolItNetworkNr["STU1"])
            self.sConfig = sConfig        

    def bSampleSetupSet(self, sSetup):
        bReturn = False
        self.sSetupConfig = sSetup
        for config in self.tree.find('Config'):
            if self.sSetupConfig == config.get('name'):
                bReturn = True
                break
        return bReturn
        
    def bLogSet(self, sLogLocation):
        bOk = False
        if -1 != sLogLocation.rfind('.'):
            sLogLocation = rreplace(sLogLocation, '.', "_" + self.sDateClock() + ".")
            logNameError = sLogLocation
            logNameError = rreplace(logNameError, '.', "bError.")
            self.PeakCan.vLogNameChange(sLogLocation, logNameError)
            bOk = True
        return bOk
    
    def vLogCountInc(self):
        fileName = self.PeakCan.Logger.fileName[:-24]
        fileName = fileName + "_" + self.sDateClock() + ".txt"
        logNameError = fileName
        logNameError = rreplace(logNameError, '.', "bError.")
        self.PeakCan.vLogNameCloseInterval(fileName, logNameError)
        
    def vSheetFileSet(self, sSheetFile):
        self.sSheetFile = sSheetFile
        
    def vAccSet(self, bX, bY, bZ, dataSets):
        self.bAccX = bool(bX) 
        self.bAccY = bool(bY) 
        self.bAccZ = bool(bZ) 

        if  (dataSets in DataSets):
            self.tAccDataFormat = DataSets[dataSets]
        else:
            dataSets = self.PeakCan.dataSetsCan20(bX, bY, bZ)
            self.tAccDataFormat = DataSets[dataSets]
        
    def vVoltageSet(self, bX, bY, bZ, dataSets):
        self.bVoltageX = bool(bX)
        self.bVoltageY = bool(bY)
        self.bVoltageZ = bool(bZ)
        
        if (dataSets in DataSets):
            self.tVoltageDataFormat = DataSets[dataSets]
        else:
            dataSets = self.PeakCan.dataSetsCan20(bX, bY, bZ)
            self.tVoltageDataFormat = DataSets[dataSets]
    
    def vSthAutoConnect(self, bSthAutoConnect):     
        self.bSthAutoConnect = bool(bSthAutoConnect)
        
    def vDeviceNameSet(self, sDevName):
        if 8 < len(sDevName):
            sDevName = sDevName[:8]
        self.sDevName = sDevName
        self.iDevNr = None
        
    def vDeviceAddressSet(self, iAddress):
        iAddress = int(iAddress, base=0)
        if 0 <= iAddress and (2 ** 48 - 1) > iAddress:
            iAddress = hex(iAddress)
            self.iAddress = iAddress
        else:
            self.iAddress = 0        
        
    def vAdcConfig(self, iPrescaler, iAquistionTime, iOversampling):
        if Prescaler["Min"] > iPrescaler:
            iPrescaler = Prescaler["Min"]
        elif Prescaler["Max"] < iPrescaler:
            iPrescaler = Prescaler["Max"]  
        try:  
            iAcquisitionTime = AdcAcquisitionTime[iAquistionTime]
            iOversampling = AdcOverSamplingRate[iOversampling]
            bTakeIt = True
        except:
            bTakeIt = False
            
        if False != bTakeIt:
            self.samplingRate = int(calcSamplingRate(iPrescaler, iAcquisitionTime, iOversampling) + 0.5)
            self.iPrescaler = iPrescaler
            self.iAquistionTime = iAcquisitionTime
            self.iOversampling = iOversampling
        
    def vAdcRefVConfig(self, sAdcRef):
        self.sAdcRef = sAdcRef
        
    def vDisplayTime(self, iDisplayTime):
        if Watch["DisplayTimeMax"] < iDisplayTime:
            iDisplayTime = Watch["DisplayTimeMax"]
        self.iDisplayTime = int(iDisplayTime) 
        
    def vRunTime(self, runTime, intervalTime):
        self.iIntervalTime = int(intervalTime)
        if Watch["IntervalDimMinX"] > self.iIntervalTime:
            self.iIntervalTime = 0
        self.iRunTime = int(runTime)

    """
    sampleInterval in ms
    """

    def vGraphInit(self, sampleInterval=200, blockSize=10):
        self.tDataPointTimeStamp = 0
        self.iPacketLossTimeStamp = 0
        self.iGraphBlockSize = blockSize
        self.iGraphSampleInterval = sampleInterval
        self.sMsgLoss = "Acceleration(" + str(format(0, '3.3f')) + "%)"
        self.GuiPackage = {"X":[], "Y" : [], "Z" : []}
        self.tSocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        
                        
    def guiProcessStop(self):
        try:
            self.vGraphSend(["Run", False])
            self.tSocket.close()     
            self.guiProcess.join()
        except:
            pass
    
    def vGraphSend(self, data):
        bSend = True
        data = tArray2Binary(data)
        while False != bSend:
            self.tSocket.sendall(data)
            sleep(0.1)
            ack = self.tSocket.recv(2**10)   
            if None != ack:
                if data == ack:
                    bSend = False
                     
    def guiProcessRestart(self):
        self.guiProcessStop()       
        if 0 < self.iDisplayTime:   
            self.guiProcess = multiprocessing.Process(target=vPlotter)
            self.guiProcess.start()
            print("Connect to GUI")
            self.tSocket.connect((HOST, PORT))
            self.vGraphSend(["dataBlockSize", self.iGraphBlockSize])
            self.vGraphSend(["sampleInterval", self.iGraphSampleInterval])
            self.vGraphSend(["xDim", self.iDisplayTime])
            self.vGraphPacketLossUpdate(0)
            if False != self.bAccX:
                self.vGraphSend(["lineNameX", "AccX"])
            if False != self.bAccY:
                self.vGraphSend(["lineNameY", "AccY"])
            if False != self.bAccZ:
                self.vGraphSend(tArray2Binary())
            self.vGraphSend(["Plot", True])
            
     
     
     
    def vGraphPointNext(self, x, y, z):
        if 0 < self.iDisplayTime:  
            if False != self.guiProcess.is_alive():
                timeStampNow = int(round(time() * 1000))
                if self.iGraphSampleInterval / self.iGraphBlockSize <= (timeStampNow - self.tDataPointTimeStamp):
                    self.tDataPointTimeStamp = timeStampNow
                    self.GuiPackage["X"].append(x)
                    self.GuiPackage["Y"].append(y)
                    self.GuiPackage["Z"].append(z)
                    if self.iGraphBlockSize <= len(self.GuiPackage["X"]):
                        try:
                            self.tSocket.sendall(tArray2Binary(["data", self.GuiPackage]))
                        except:
                            pass
                        self.GuiPackage = {"X":[], "Y" : [], "Z" : []}
            else:
                self.aquireEndTime = self.PeakCan.getTimeMs()

    def vGraphPacketLossUpdate(self, msgCounter):
        if 0 < self.iDisplayTime:  
            self.iMsgCounterLast += 1
            self.iMsgCounterLast %= 256
            if  self.iMsgCounterLast != msgCounter:
                iLost = msgCounter - self.iMsgCounterLast
                self.iMsgLoss += iLost
                self.iMsgsTotal += iLost
                if 0 > iLost:
                    self.iMsgLoss += 256
                    self.iMsgsTotal += 256
                self.iMsgCounterLast = msgCounter
            else:
                self.iMsgsTotal += 1
        
        iPacketLossTimeStamp = int(round(time() * 1000))
        if 1000 <= (iPacketLossTimeStamp - self.iPacketLossTimeStamp):
            self.iPacketLossTimeStamp = iPacketLossTimeStamp  
            sMsgLoss = "Acceleration(" + str(format(100-(100 * self.iMsgLoss / self.iMsgsTotal), '3.3f')) + "%)"     
            if sMsgLoss != self.sMsgLoss:
                self.sMsgLoss = sMsgLoss
                self.tSocket.sendall(tArray2Binary(["diagramName", self.sMsgLoss]))                
            self.iMsgLoss = 0
            self.iMsgsTotal = 0
        
    def vVersion(self, major, minor, build):
        if 2 <= major and 1 <= minor:
            self.Major = major
            self.Minor = minor
            self.Build = build
            
    def sDateClock(self):
        DataClockTimeStamp = datetime.fromtimestamp(time()).strftime('%Y-%m-%d_%H-%M-%S')
        return DataClockTimeStamp

    def vParserInit(self):
        self.parser = argparse.ArgumentParser(description='Command Line Oprtions')
        self.parser.add_argument('-a', '--adc', dest='adc_config', action='store', nargs=3, type=int, required=False, help='Prescaler AcquisitionTime OversamplingRate (3 inputs required in that order e.g. 2 8 64) - Note that acceleration and battery voltage measurements share a single ADC that samples up to 4 channels)')
        self.parser.add_argument('-b', '--bluetooth_connect', dest='bluetooth_connect', action='store', nargs=1, type=str, required=False, help='Connect to device specified by Bluetooth address and starts sampling as configured')
        self.parser.add_argument('-c', '--auto_connect', dest='auto_connect', action='store_true', required=False, help='Automatically connect to Device as defined in the given xml file')
        self.parser.add_argument('-d', '--devs', dest='devNameList', action='store_true', required=False, help='Get Device Names, Bluetooth address and Receive Signal Strength Indicators(RSSI) of all available STHs')    
        self.parser.add_argument('-e', '--xlsx', dest='xlsx', action='store', nargs=1, type=str, required=False, help='Table Calculation File(xlsx) to transfer configuration from/to STH/STU')
        self.parser.add_argument('-i', '--interval', dest='interval', action='store', nargs=1, type=int, required=False, help='Sets Interval Time (Output file is saved each interval time in seconds. Lower than 10 causes a single file')
        self.parser.add_argument('-l', '--log_location', dest='log_name', action='store', nargs=1, type=str, required=False, help='Where to save Log Files (relative/absolute path+file name)')
        self.parser.add_argument('-n', '--name_connect', dest='name_connect', action='store', nargs=1, type=str, required=False, help='Connect to device specified by Name and starts sampling as configured')
        self.parser.add_argument('-p', '--points', dest='points', action='store', nargs=1, type=int, required=False, help='PPP specifies which acceleration axis(X/Y/Z) are active(1) or off(0)')
        self.parser.add_argument('-r', '--run_time', dest='run_time', action='store', nargs=1, type=int, required=False, help='Sets RunTime in seconds. 0 specifies infinity')
        self.parser.add_argument('-s', '--sample_setup', dest='sample_setup', action='store', nargs=1, type=str, required=False, help='Starts sampling with configuration as given including additional command line arguments')
        self.parser.add_argument('-v', '--version', dest='version', action='store', nargs=2, type=str, required=False, help='Chooses product with version for handling Table Calculation Files (e.g. STH v2.1.2)')
        self.parser.add_argument('-x', '--xml', dest='xml_file_name', action='store', nargs=1, type=str, required=True, help='Selects xml configuration/data base file')
        self.parser.add_argument('--create', dest='create', action='store_true', required=False, help='Creates a device configuration or sample setup in the xml file')
        self.parser.add_argument('--gui_dim', dest='gui_dim', action='store', nargs=1, type=int, required=False, help='Length of visualization interval in ms for the graphical acceleration view . Value below 10 turns it off')
        self.parser.add_argument('--refv', dest='refv', action='store', nargs=1, type=str, required=False, help='ADC\'s Reference voltage, VDD=Standard')
        self.parser.add_argument('--remove', dest='remove', action='store_true', required=False, help='Removes a device configuration or sample setup in the xml file')
        self.parser.add_argument('--save', dest='save', action='store_true', required=False, help='Saves a device configuration or sample setup in the xml file)')
        self.parser.add_argument('--show_config', dest='show_config', action='store_true', required=False, help='Shows current configuration (including command line arguments)')
        self.parser.add_argument('--show_products', dest='show_products', action='store_true', required=False, help='Shows all available devices and additional versions')
        self.parser.add_argument('--show_setups', dest='show_setups', action='store_true', required=False, help='Shows current configuration (including command line arguments)')
        self.parser.add_argument('--voltage_points', dest='voltage_points', action='store', nargs=1, type=int, required=False, help='PPP specifies which voltage axis (sample points; X/Y/Z) are active(1) or off(0). Note that x specifies the battery')
        args = self.parser.parse_args()
        self.args_dict = vars(args)
    
    def vParserConsoleArgumentsPassXml(self):
        if None != self.args_dict['version'] and None != self.args_dict['sample_setup']:
            print("You can't use sample setup and product/version simultaneously")
            self.PeakCan.vLogDel()
            self.__exit__()
        bRemove = False  
        if False != self.args_dict['remove']:
            bRemove = True      
            self.vSave2Xml(True)          
        bCreate = False        
        if False != self.args_dict['create']:
            if False != bRemove:
                print("You can't create and remove simultaneously.")
                self.PeakCan.vLogDel()
                self.__exit__()
            else:
                bCreate = True     
                self.vSave2Xml(True)
                       
        self.vConfigFileSet(self.args_dict['xml_file_name'][0])
                  
        if False != self.args_dict['save']:
            self.vSave2Xml(True)      
        if None != self.args_dict['sample_setup']: 
            sSetup = self.args_dict['sample_setup'][0]
            bSetupFound = self.bSampleSetupSet(sSetup)
            if False != bCreate:
                if False != bSetupFound:
                    print("Sample Configuration not found")
                    self.__exit_()
                else:
                    self.newXmlSetup(sSetup)
                    self.vSave2Xml(True)
                    bCreate = False
            elif False != bRemove:
                if False != bSetupFound: 
                    if 1 < len(self.tree.find('Config')):
                        self.removeXmlSetup(sSetup)
                    else:
                        print("You cant remove the last sample setup")
                        self.__exit_()
                else:
                    print("You try to remove something that does not exist")
                    self.__exit_()
            else:
                self.vGetXmlSetup()
                
        if None != self.args_dict['version']: 
            self.vConfigSet(self.args_dict['version'][0], self.args_dict['version'][1])
            if False != bCreate or False != bRemove:
                dataDef = self.root.find('Data')
                for product in dataDef.find('Product'):
                    if product.get('name') == self.sProduct:
                        break
                if product.get('name') == self.sProduct:
                    for productVersion in product.find('Version'):
                        if productVersion.get('name') == self.sConfig:
                            break
                    if productVersion.get('name') != self.sConfig and False != bCreate:
                        self.newXmlVersion(product, product.find('Version')[0], self.sConfig)
                        self.vSave2Xml(True)
                    elif productVersion.get('name') == self.sConfig and False != bRemove:
                        if 1 < len(product.find('Version')):
                            self.removeXmlVersion(product.find('Version'), productVersion)
                    elif False != bCreate:
                        print("Error! you tried to create a product/version that allready exists")
                        self.PeakCan.vLogDel()
                        self.__exit__()
                    else:
                        print("Error! you tried to create a sample setup that allready exists")
                        self.PeakCan.vLogDel()
                        self.__exit__()
        
    def vParserConsoleArgumentsPass(self):  
        self.vParserConsoleArgumentsPassXml()    
        if None != self.args_dict['gui_dim']:
            self.vDisplayTime(self.args_dict['gui_dim'][0])            
        if None != self.args_dict['log_name']:
            self.bLogSet(self.args_dict['log_name'][0]) 
        if None != self.args_dict['adc_config']:
            adcConfig = self.args_dict['adc_config']
            self.vAdcConfig(adcConfig[0], adcConfig[1], adcConfig[2])
        if None != self.args_dict['refv']:
            self.vAdcRefVConfig(self.args_dict['refv'][0])
        if None != self.args_dict['xlsx']:
            self.vSheetFileSet(self.args_dict['xlsx'][0])
        iIntervalTime = self.iIntervalTime
        if None != self.args_dict['interval']:
            iIntervalTime = self.args_dict['interval'][0]           
        iRunTime = self.iRunTime
        if None != self.args_dict['run_time']:
            iRunTime = self.args_dict['run_time'][0]
        self.vRunTime(iRunTime, iIntervalTime)

        if None != self.args_dict['name_connect']:
            self.vDeviceNameSet(self.args_dict['name_connect'][0])
            self.vSthAutoConnect(True)
        elif None != self.args_dict['bluetooth_connect']:
            self.vDeviceAddressSet(self.args_dict['bluetooth_connect'][0])
            self.vSthAutoConnect(True)        
        elif False != self.args_dict['auto_connect']:    
            self.vSthAutoConnect(True)   
            
        if None != self.args_dict['points']: 
            points = self.args_dict['points'][0] & 0x07
            bZ = bool(points & 1)
            bY = bool((points >> 1) & 1)
            bX = bool((points >> 2) & 1)
            self.vAccSet(bX, bY, bZ, -1)
        if None != self.args_dict['voltage_points']: 
            points = self.args_dict['voltage_points'][0] & 0x07
            bZ = bool(points & 1)
            bY = bool((points >> 1) & 1)
            bX = bool((points >> 2) & 1)
            self.vVoltageSet(bX, bY, bZ, -1)                          
                           
    def reset(self):
        if False == self.KeyBoadInterrupt:
            try:
                self.PeakCan.ReadThreadReset()
                self.PeakCan.cmdReset(MyToolItNetworkNr["STU1"])
            except KeyboardInterrupt:
                self.KeyBoadInterrupt = True

    def BlueToothConnect(self):
        try:
            self.iDevNr = int(input('Input:'))  
            try:
                self.PeakCan.BlueToothConnect(MyToolItNetworkNr["STU1"], self.iDevNr)          
            except KeyboardInterrupt:
                self.KeyBoadInterrupt = True
        except ValueError:
            print("Not a number") 
    
    def getBlueToothDeviceList(self): 
        self.PeakCan.BlueToothConnectConnect(MyToolItNetworkNr["STH1"])
        deviceNumbers = 0
        endTime = time() + BlueToothDeviceListAquireTime
        while(time() < endTime):
            deviceNumbers = self.PeakCan.BlueToothConnectTotalScannedDeviceNr(MyToolItNetworkNr["STH1"])
        nameList = []
        for dev in range(deviceNumbers):
            nameList.append([dev, self.PeakCan.BlueToothNameGet(dev)])
        return nameList   
                
    def BlueToothConnectName(self):
        if False == self.KeyBoadInterrupt:
            try:
                self.PeakCan.BlueToothConnectPollingName(MyToolItNetworkNr["STU1"], self.sDevName)
            except KeyboardInterrupt:
                self.KeyBoadInterrupt = True
                self.__exit__()
               
    def vDataAquisition(self):  
        if False == self.KeyBoadInterrupt:
            try:
                if False != self.PeakCan.bConnected: 
                    self.PeakCan.ConfigAdc(MyToolItNetworkNr["STH1"], self.iPrescaler, self.iAquistionTime, self.iOversampling, AdcReference[self.sAdcRef])
                    self.PeakCan.readThreadStop()     
                    self.guiProcessRestart()       
                    print("Start")
                    self.PeakCan.Logger.Info("Start")
                    self.vGetStreamingAccData()
                    self.guiProcessStop()
                else:
                    print("Device not allocable")    
                    self.PeakCan.Logger.bError("Device not allocable")     
            except KeyboardInterrupt:
                self.KeyBoadInterrupt = True
                self.__exit__()
                
    def close(self):          
        if False != self.PeakCan.RunReadThread:
            self.__exit__()  

    def vGetStreamingAccDataProcess(self):
        iIntervalTime = self.iIntervalTime * 1000
        if 0 == self.iIntervalTime:
            iIntervalTime += (1 << 32)
        startTime = self.PeakCan.getTimeMs()
        try:
            while(self.PeakCan.getTimeMs() < self.aquireEndTime):
                if (self.PeakCan.getTimeMs() - startTime) >= iIntervalTime:
                    startTime = self.PeakCan.getTimeMs()
                    self.vLogCountInc()
                ack = self.ReadMessage()
                if(None != ack):
                    if(self.AccAckExpected.ID != ack["CanMsg"].ID and self.VoltageAckExpected.ID != ack["CanMsg"].ID):
                        self.PeakCan.Logger.bError("CanId bError: " + str(ack["CanMsg"].ID))
                    elif(self.AccAckExpected.DATA[0] != ack["CanMsg"].DATA[0] and self.VoltageAckExpected.DATA[0] != ack["CanMsg"].DATA[0])  :
                        self.PeakCan.Logger.bError("Wrong Subheader-Format(Acceleration Format): " + str(ack["CanMsg"].ID))
                    elif self.AccAckExpected.ID == ack["CanMsg"].ID:
                        self.GetMessageAcc(ack)
                                   
                    else:
                        self.GetMessageVoltage(ack)     
        except KeyboardInterrupt:
            self.KeyBoadInterrupt = True
            print("Data acquisition determined")
            self.__exit__()                 
    
    def vGetStreamingAccDataAccStart(self): 
        accFormat = AtvcFormat()
        accFormat.asbyte = 0
        accFormat.b.bStreaming = 1
        accFormat.b.bNumber1 = self.bAccX
        accFormat.b.bNumber2 = self.bAccY
        accFormat.b.bNumber3 = self.bAccZ
        accFormat.b.u3DataSets = self.tAccDataFormat
        cmd = self.PeakCan.CanCmd(MyToolItBlock["Streaming"], MyToolItStreaming["Acceleration"], 0, 0)
        self.AccAckExpected = self.PeakCan.CanMessage20(cmd, MyToolItNetworkNr["STH1"], MyToolItNetworkNr["SPU1"], [accFormat.asbyte])
        cmd = self.PeakCan.CanCmd(MyToolItBlock["Streaming"], MyToolItStreaming["Acceleration"], 1, 0)
        message = self.PeakCan.CanMessage20(cmd, MyToolItNetworkNr["SPU1"], MyToolItNetworkNr["STH1"], [accFormat.asbyte])
        self.PeakCan.Logger.Info("MsgId/Subpayload(Acc): " + hex(message.ID) + "/" + hex(accFormat.asbyte))
        ack = None
        endTime = self.PeakCan.getTimeMs() + 4000
        while (None == ack) and (self.PeakCan.getTimeMs() < endTime):
            self.PeakCan.WriteFrame(message)
            readEndTime = self.PeakCan.getTimeMs() + 500
            while((None == ack) and  (self.PeakCan.getTimeMs() < readEndTime)):
                ack = self.ReadMessage()        
        return ack

    def vGetStreamingAccDataVoltageStart(self): 
        voltageFormat = AtvcFormat()
        voltageFormat.asbyte = 0
        voltageFormat.b.bStreaming = 1
        voltageFormat.b.bNumber1 = self.bVoltageX
        voltageFormat.b.bNumber2 = self.bVoltageY
        voltageFormat.b.bNumber3 = self.bVoltageZ
        voltageFormat.b.u3DataSets = self.tVoltageDataFormat
        cmd = self.PeakCan.CanCmd(MyToolItBlock["Streaming"], MyToolItStreaming["Voltage"], 0, 0)
        self.VoltageAckExpected = self.PeakCan.CanMessage20(cmd, MyToolItNetworkNr["STH1"], MyToolItNetworkNr["SPU1"], [voltageFormat.asbyte])
        cmd = self.PeakCan.CanCmd(MyToolItBlock["Streaming"], MyToolItStreaming["Voltage"], 1, 0)
        message = self.PeakCan.CanMessage20(cmd, MyToolItNetworkNr["SPU1"], MyToolItNetworkNr["STH1"], [voltageFormat.asbyte])
        self.PeakCan.Logger.Info("MsgId/Subpayload(Voltage): " + hex(message.ID) + "/" + hex(voltageFormat.asbyte))
        ack = None
        endTime = self.PeakCan.getTimeMs() + 4000
        while (None == ack) and (self.PeakCan.getTimeMs() < endTime):
            self.PeakCan.WriteFrame(message)
            readEndTime = self.PeakCan.getTimeMs() + 500
            while((None == ack) and  (self.PeakCan.getTimeMs() < readEndTime)):
                ack = self.ReadMessage()        
        return ack
                                 
    def vGetStreamingAccData(self):  
        ack = self.vGetStreamingAccDataAccStart()
        if None != ack:
            ack = self.vGetStreamingAccDataVoltageStart()
        currentTime = self.PeakCan.getTimeMs()
        if None == ack:
            self.PeakCan.Logger.bError("No Ack received from Device: " + str(self.iDevNr))
            self.aquireEndTime = currentTime
        elif(0 == self.iRunTime):
            self.aquireEndTime = currentTime + (1 << 32)
        else:
            self.aquireEndTime = currentTime + self.iRunTime * 1000
        self.vGetStreamingAccDataProcess()
        
    def GetMessageSingle(self, prefix, canMsg):  
        canData = canMsg["CanMsg"].DATA 
        p1 = messageValueGet(canData[2:4])
        p2 = messageValueGet(canData[6:8])
        p3 = messageValueGet(canData[4:6])   
        
        canTimeStamp = canMsg["PeakCanTime"]
        canTimeStamp = round(canTimeStamp, 3)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix + " ")
        ackMsg += str(format(p1, '5d'))
        ackMsg += "; "
        self.PeakCan.Logger.Info(ackMsg)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix + " ")
        ackMsg += str(format(p2, '5d'))
        ackMsg += "; "
        self.PeakCan.Logger.Info(ackMsg)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix + " ")
        ackMsg += str(format(p3, '5d'))
        ackMsg += "; "
        self.PeakCan.Logger.Info(ackMsg)  
        
    def GetMessageDouble(self, prefix1, prefix2, canMsg):
        canData = canMsg["CanMsg"].DATA
        canTimeStamp = canMsg["PeakCanTime"]
        canTimeStamp = round(canTimeStamp, 3)
        p1_1 = messageValueGet(canData[2:4])
        p1_2 = messageValueGet(canData[4:6])
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix1 + " ")
        ackMsg += str(format(p1_1, '5d'))
        ackMsg += "; "
        ackMsg += prefix2
        ackMsg += " "
        ackMsg += str(format(p1_2, '5d'))
        ackMsg += "; "
        self.PeakCan.Logger.Info(ackMsg) 

    def GetMessageTripple(self, prefix1, prefix2, prefix3, canMsg):
        canData = canMsg["CanMsg"].DATA
        canTimeStamp = canMsg["PeakCanTime"]
        canTimeStamp = round(canTimeStamp, 3)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix1 + " ")
        ackMsg += str(format(messageValueGet(canData[2:4]), '5d'))
        ackMsg += "; "
        ackMsg += prefix2
        ackMsg += " "
        ackMsg += str(format(messageValueGet(canData[4:6]), '5d'))
        ackMsg += "; "
        ackMsg += prefix3
        ackMsg += " "
        ackMsg += str(format(messageValueGet(canData[6:8]), '5d'))
        ackMsg += "; "
        self.PeakCan.Logger.Info(ackMsg)                        

    def GetMessageAcc(self, canData):
        data = canData["CanMsg"].DATA
        msgCounter = data[1]
        self.vGraphPacketLossUpdate(msgCounter)
        
        if self.tAccDataFormat == DataSets[1]:
            if (False != self.bAccX) and (False != self.bAccY) and (False == self.bAccZ):
                self.GetMessageDouble("AccX", "AccY", canData)
                self.vGraphPointNext(messageValueGet(data[2:4]), messageValueGet(data[4:6]), 0)
            elif (False != self.bAccX) and (False == self.bAccY) and (False != self.bAccZ):
                self.GetMessageDouble("AccX", "AccZ", canData)
                self.vGraphPointNext(messageValueGet(data[2:4]), 0, messageValueGet(data[4:6]))
            elif (False == self.bAccX) and (False != self.bAccY) and (False != self.bAccZ):
                self.GetMessageDouble("AccY", "AccZ", canData) 
                self.vGraphPointNext(0, messageValueGet(data[2:4]), messageValueGet(data[4:6]))
            else:
                self.GetMessageTripple("AccX", "AccY", "AccZ", canData)   
                self.vGraphPointNext(messageValueGet(data[2:4]), messageValueGet(data[4:6]), messageValueGet(data[6:8]))
        elif self.tAccDataFormat == DataSets[3]:
            if False != self.bAccX:
                self.GetMessageSingle("AccX", canData)  
                self.vGraphPointNext(messageValueGet(data[2:4]), 0, 0)            
            elif False != self.bAccY:
                self.GetMessageSingle("AccY", canData)   
                self.vGraphPointNext(0, messageValueGet(data[2:4]), 0)          
            elif False != self.bAccZ:
                self.GetMessageSingle("AccZ", canData) 
                self.vGraphPointNext(0, 0, messageValueGet(data[2:4]))   
        else:               
            self.PeakCan.Logger.bError("Wrong Ack format")

    def GetMessageVoltage(self, canData):
        if self.tAccDataFormat == DataSets[1]:
            if (0 != self.bVoltageX) and (0 != self.bVoltageY) and (0 == self.bVoltageZ):
                self.GetMessageDouble("VoltageX", "VoltageY", canData)
            elif (0 != self.bVoltageX) and (0 == self.bVoltageY) and (0 != self.bVoltageZ):
                self.GetMessageDouble("VoltageX", "VoltageZ", canData)
            elif (0 == self.bVoltageX) and (0 != self.bVoltageY) and (0 != self.bVoltageZ):
                self.GetMessageDouble("VoltageY", "VoltageZ", canData) 
            else:
                self.GetMessageTripple("VoltageX", "VoltageY", "VoltageZ", canData)   
        elif self.tVoltageDataFormat == DataSets[3]:
            if 0 != self.bVoltageX:
                self.GetMessageSingle("VoltageX", canData)               
            elif 0 != self.bVoltageY:
                self.GetMessageSingle("VoltageY", canData)               
            elif 0 != self.bVoltageZ:
                self.GetMessageSingle("VoltageZ", canData)       
        else:               
            self.PeakCan.Logger.bError("Wrong Ack format")
            
    def ReadMessage(self):
        message = None
        result = self.PeakCan.m_objPCANBasic.Read(self.PeakCan.m_PcanHandle)
        if result[0] == PeakCanFd.PCAN_ERROR_OK:
            peakCanTimeStamp = result[2].millis_overflow * (2 ** 32) + result[2].millis + result[2].micros / 1000
            message = {"CanMsg" : result[1], "PcTime" : self.PeakCan.getTimeMs(), "PeakCanTime" : peakCanTimeStamp}   
        elif result[0] == PeakCanFd.PCAN_ERROR_QOVERRUN:
            self.Logger.bError("RxOverRun")
            print("RxOverRun")
            raise
        else:
            sleep(0.0002)
        return message


    def atXmlProductVersion(self):
        dataDef = self.root.find('Data')               
        asProducts = {}
        iProduct = 1
        for product in dataDef.find('Product'):
            asProducts[iProduct] = {}
            asProducts[iProduct]["Product"] = product
            asProducts[iProduct]["Versions"] = {}            
            iVersion = 1
            for version in product.find('Version'):
                asProducts[iProduct]["Versions"][iVersion] = version
                iVersion += 1      
            iProduct+=1          
        return asProducts    
    
    def excelCellWidthAdjust(self, worksheet, factor=1.2, bSmaller=True):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                    continue
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * factor
            columLetter = chr(ord('A') + column - 1)
            if adjusted_width > worksheet.column_dimensions[columLetter].width or False == bSmaller:
                worksheet.column_dimensions[columLetter].width = adjusted_width
            
    """
    Create Excel Sheet by xml definition
    """

    def excelSheetCreate(self):
        dataDef = self.root.find('Data')
        for product in dataDef.find('Product'):
            if product.get('name') == self.sProduct:
                break
        for version in product.find('Version'):
            if version.get('name') == self.sConfig:
                break
        if version.get('name') == self.sConfig:
            workbook = openpyxl.Workbook()
            FontRow1 = Font(bold=True, size=20)
            FontRowRow2 = Font(bold=False, size=12)
            workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
            
            for page in version.find('Page'):
                i = 2
                name = page.get('name')
                pageAddress = int(page.find('pageAddress').text)
                worksheet = workbook.create_sheet(name + "@" + hex(pageAddress))
                worksheet['A1'] = 'Name'
                worksheet['A1'].font = FontRow1
                worksheet['B1'] = 'Address'
                worksheet['B1'].font = FontRow1
                worksheet['C1'] = 'Length'
                worksheet['C1'].font = FontRow1
                worksheet['D1'] = 'Read Only'
                worksheet['D1'].font = FontRow1
                worksheet['E1'] = 'Value'
                worksheet['E1'].font = FontRow1
                worksheet['F1'] = 'Unit'
                worksheet['F1'].font = FontRow1
                worksheet['G1'] = 'Format'
                worksheet['G1'].font = FontRow1
                worksheet['H1'] = 'Description'
                worksheet['H1'].font = FontRow1
                self.excelCellWidthAdjust(worksheet, 1.6, False)
                for entry in page.find('Entry'):
                    worksheet['A' + str(i)] = entry.get('name')
                    worksheet['A' + str(i)].font = FontRowRow2
                    worksheet['B' + str(i)] = int(entry.find('subAddress').text)
                    worksheet['B' + str(i)].font = FontRowRow2
                    worksheet['C' + str(i)] = int(entry.find('length').text)
                    worksheet['C' + str(i)].font = FontRowRow2
                    worksheet['D' + str(i)] = entry.find('readOnly').text
                    worksheet['D' + str(i)].font = FontRowRow2
                    try:
                        worksheet['E' + str(i)] = int(entry.find('value').text)
                    except ValueError:
                        worksheet['E' + str(i)] = entry.find('value').text
                    worksheet['E' + str(i)].font = FontRowRow2
                    worksheet['F' + str(i)] = entry.find('unit').text
                    worksheet['F' + str(i)].font = FontRowRow2
                    worksheet['G' + str(i)] = entry.find('format').text
                    worksheet['G' + str(i)].font = FontRowRow2
                    worksheet['H' + str(i)] = entry.find('description').text
                    worksheet['H' + str(i)].font = FontRowRow2
                    i += 1
                self.excelCellWidthAdjust(worksheet)
            workbook.save(self.sSheetFile)
    
    def _excelSheetEntryFind(self, entry, key, value):
        if None != value:
            entry.find(key).text = str(value)

    """
    Set endoding
    """

    def _XmlWriteEndoding(self):
        xml = (bytes('<?xml version="1.0" encoding="UTF-8"?>\n', encoding='utf-8') + ET.tostring(self.root))
        xml = xml.decode('utf-8')
        with open(self.sXmlFileName, "w", encoding='utf-8') as f:
            f.write(xml)   
     
    """
    Creats a new config
    """

    def newXmlVersion(self, product, productVersion, sVersion):
        cloneVersion = copy.deepcopy(productVersion)
        cloneVersion.set('name', sVersion) 
        product.find('Version').append(cloneVersion)        
        self.xmlSave()    
        self.vConfigSet(product.get('name'), sVersion)

    """
    Save XML File (in any state)
    """

    def xmlSave(self):
        self.tree.write(self.sXmlFileName)
        self._XmlWriteEndoding()   
        del self.tree
        self.vConfigFileSet(self.sXmlFileName)
        
    """
    Removes a config
    """

    def removeXmlVersion(self, product, vesion):
        product.find('Version').remove(vesion)
        self.xmlSave()

    def xmlPrintVersions(self):
        dataDef = self.root.find('Data')
        for product in dataDef.find('Product'):
            print(product.get('name') + ":")
            for version in product.find('Version'):
                print("   " + version.get('name'))
        
    """
    Write xml definiton by Excel Sheet
    """

    def excelSheetConfig(self):
        dataDef = self.root.find('Data')
        for product in dataDef.find('Product'):
            if product.get('name') == self.sProduct:
                break
        for version in product.find('Version'):
            if version.get('name') == self.sConfig:
                break
        workbook = openpyxl.load_workbook(self.sSheetFile)
        if workbook:
            if version.get('name') != self.sConfig:
                self.newXmlVersion(product, product.find('Version')[0], self.sConfig)
                self.xmlSave()
                self.excelSheetConfig()
            else:
                for worksheetName in workbook.sheetnames:
                    name = str(worksheetName).split('@')
                    address = name[1]
                    name = name[0]
                    for page in version.find('Page'):
                        i = 2
                        pageName = page.get('name')
                        pageAddress = hex(int(page.find('pageAddress').text))
                        if name == pageName and pageAddress == address:
                            worksheet = workbook.get_sheet_by_name(worksheetName)
                            for entry in page.find('Entry'):
                                value = str(worksheet['A' + str(i)].value)
                                if None != value:
                                    entry.set('name', value)
                                else:
                                    entry.set('name', "")
                                self._excelSheetEntryFind(entry, 'subAddress', worksheet['B' + str(i)].value)
                                self._excelSheetEntryFind(entry, 'length', worksheet['C' + str(i)].value)
                                self._excelSheetEntryFind(entry, 'readOnly', worksheet['D' + str(i)].value)
                                self._excelSheetEntryFind(entry, 'value', worksheet['E' + str(i)].value)
                                self._excelSheetEntryFind(entry, 'unit', worksheet['F' + str(i)].value)
                                self._excelSheetEntryFind(entry, 'format', worksheet['G' + str(i)].value)
                                self._excelSheetEntryFind(entry, 'description', worksheet['H' + str(i)].value)
                                i += 1
                self.xmlSave()
    
    
    def iExcelSheetPageLength(self, worksheet):
        i = 2
        totalLength = 0
        while 256 > i:
            if None != worksheet['A' + str(i)].value:
                length = int(worksheet['C' + str(i)].value)
                totalLength += length
                i+=length
            else:
                break
        return totalLength 
    
    def vExcelSheetPageValue(self, worksheet, byteArry):
        i = 2
        totalLength = 0
        while len(byteArry) > totalLength:
            if None != worksheet['A' + str(i)].value:
                length = worksheet['C' + str(i)].value
                worksheet['E' + str(i)] = array2Value(byteArry[totalLength:], length)
                totalLength += length
                i+=length
            else:
                break
        return totalLength    
    
    
    def atExcelSheetNames(self):
        workSheetNames = []
        workbook = openpyxl.load_workbook(self.sSheetFile)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = name[1]
                name = name[0]
                workSheetNames.append(name)
        return workSheetNames
    
    """
    Read EEPROM page to write values in Excel Sheet
    """    

    def excelSheetRead(self, namePage):
        workbook = openpyxl.load_workbook(self.sSheetFile)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = name[1]
                name = name[0]
                if namePage == name:
                    worksheet = workbook.get_sheet_by_name(worksheetName)
                    pageContent = []
                    readLength = self.iExcelSheetPageLength(worksheet)
                    for offset in range(0, readLength, 4):
                        index = self.PeakCan.cmdSend(MyToolItNetworkNr["STH1"], MyToolItBlock["Eeprom"], MyToolItEeprom["Read"], [address, 0xFF & offset, 4, 0, 0, 0, 0, 0])   
                        pageContent.extend(self.PeakCan.getReadMessageData(index)[4:])
                    pageContent = pageContent[0:readLength]
                    self.vExcelSheetPageValue(worksheet, pageContent)
            workbook.save(self.sSheetFile)    
                
                
                
    """
    Write EEPROM to write values in Excel Sheet
    """    

    def excelSheetWrite(self):
        pass
    
    
    def atXmlSetup(self):           
        asSetups = {}
        iSetup = 1
        for setup in self.tree.find('Config'):
            asSetups[iSetup] = setup      
            iSetup += 1   
        return asSetups  
    
    
    def vSetXmlSetup(self):
        for config in self.tree.find('Config'):
            if config.get('name') == self.sSetupConfig:        
                config.find('DeviceName').text = str(self.sDevName)
                config.find('DeviceAddress').text = str(self.iAddress)
                config.find('Acc').text = str(int(self.bAccX)) + str(int(self.bAccY)) + str(int(self.bAccZ))
                config.find('Voltage').text = str(int(self.bVoltageX)) + str(int(self.bVoltageY)) + str(int(self.bVoltageZ))
                config.find('Prescaler').text = str(self.iPrescaler)
                config.find('AcquisitionTime').text = str(AdcAcquisitionTimeReverse[self.iAquistionTime])
                config.find('OverSamples').text = str(AdcOverSamplingRateReverse[self.iOversampling])
                config.find('AdcRef').text = str(self.sAdcRef)
                sFileName = self.PeakCan.Logger.fileName
                config.find('LogName').text = sFileName[:sFileName.find('_'):]
                config.find('RunTime').text = str(self.iRunTime)
                config.find('IntervalTime').text = str(self.iIntervalTime)
                config.find('DisplayTime').text = str(self.iDisplayTime)
                break
                            
    def vGetXmlSetup(self):
        for config in self.tree.find('Config'):
            if config.get('name') == self.sSetupConfig:
                self.vDeviceNameSet(config.find('DeviceName').text)
                self.vDeviceAddressSet(config.find('DeviceAddress').text)
                samplePoints = config.find('Acc').text
                bAccX = int(samplePoints[0])
                bAccY = int(samplePoints[1])
                bAccZ = int(samplePoints[2])
                self.vAccSet(bAccX, bAccY, bAccZ, -1)
                samplePoints = config.find('Voltage').text
                bVoltageX = int(samplePoints[0])
                bVoltageY = int(samplePoints[1])
                bVoltageZ = int(samplePoints[2])
                self.vVoltageSet(bVoltageX, bVoltageY, bVoltageZ, -1)
                self.vAdcConfig(int(config.find('Prescaler').text), int(config.find('AcquisitionTime').text), int(config.find('OverSamples').text))
                self.vAdcRefVConfig(config.find('AdcRef').text)
                self.bLogSet(str(config.find('LogName').text) + ".txt")
                self.vRunTime(int(config.find('RunTime').text), int(config.find('IntervalTime').text))
                self.vDisplayTime(int(config.find('DisplayTime').text))
                break
                           
    def removeXmlSetup(self, setup):
        if( setup.get('name') == self.sSetupConfig):
            self.bSampleSetupSet(None)
        self.tree.find('Config').remove(setup)  
        self.xmlSave()     

                
    def newXmlSetup(self, setup, sConfig):
        cloneVersion = copy.deepcopy(setup)
        cloneVersion.set('name', sConfig) 
        self.tree.find('Config').append(cloneVersion)
        self.xmlSave()
        self.bSampleSetupSet(sConfig)

    def xmlPrintSetups(self):
        for setup in self.tree.find('Config'):
            print(setup.get('name'))
            print("    Device Name: " + setup.find('DeviceName').text)
            print("    Acc: " + setup.find('Acc').text)
            iAcquisitionTime = AdcAcquisitionTime[int(setup.find('AcquisitionTime').text)]
            iOversampling = AdcOverSamplingRate[int(setup.find('OverSamples').text)]
            samplingRate = int(calcSamplingRate(int(setup.find('Prescaler').text), iAcquisitionTime, iOversampling) + 0.5)
            print("    ADC Prescaler/AcquisitionTime/OversamplingRate(Samples/s): " + setup.find('Prescaler').text + "/" + setup.find('AcquisitionTime').text + "/" + setup.find('OverSamples').text + "(" + str(samplingRate) + ")")
            print("    ADC Reference Voltage: " + setup.find('AdcRef').text)
            print("    Log Name: " + setup.find('LogName').text)
            print("    RunTime/IntervalTime: " + setup.find('RunTime').text + " " + setup.find('DisplayTime').text)
            print("    Display Time: " + setup.find('DisplayTime').text)
                
    def _vRunConsoleStartupShow(self):
        print("XML File: " + str(self.sXmlFileName))
        print("Product Configuration: " + str(self.sProduct) + " " + str(self.sConfig))
        print("Setup Configuration: " + str(self.sSetupConfig))
        print("AutoSave?: " + str(self.bSave))
        print("Table Calculation File: " + str(self.sSheetFile))
        print("Log Name: " + str(self.PeakCan.Logger.fileName))
        print("Device Name (to be connected): " + str(self.sDevName))
        print("Bluetooth address(to be connected): " + str(self.iAddress))  # Todo machen
        print("AutoConnect?: " + str(self.bSthAutoConnect))
        print("Run Time: " + str(self.iRunTime) + "s")
        print("Inteval Time: " + str(self.iIntervalTime) + "s")
        print("Display Time: " + str(self.iDisplayTime) + "s")
        print("Adc Prescaler/AcquisitionTime/OversamplingRate/Reference(Samples/s): " + str(self.iPrescaler) + "/" + str(AdcAcquisitionTimeReverse[self.iAquistionTime]) + "/" + str(AdcOverSamplingRateReverse[self.iOversampling]) + "/" + str(self.sAdcRef) + "(" + str(self.samplingRate) + ")")
        print("Acc Config(XYZ/DataSets): " + str(int(self.bAccX)) + str(int(self.bAccY)) + str(int(self.bAccZ)) + "/" + str(DataSetsReverse[self.tAccDataFormat]))
        print("Voltage Config(XYZ/DataSets): " + str(int(self.bVoltageX)) + str(int(self.bVoltageY)) + str(int(self.bVoltageZ)) + "/" + str(DataSetsReverse[self.tAccDataFormat]) + ("(X=Battery)"))
        
    def _vRunConsoleStartupLoggerPrint(self):
        self.PeakCan.Logger.Info("XML File: " + str(self.sXmlFileName))
        self.PeakCan.Logger.Info("Product Configuration: " + str(self.sProduct) + " " + str(self.sConfig))
        self.PeakCan.Logger.Info("Setup Configuration: " + str(self.sSetupConfig))
        self.PeakCan.Logger.Info("AutoSave?: " + str(self.bSave))
        self.PeakCan.Logger.Info("Table Calculation File: " + str(self.sSheetFile))
        self.PeakCan.Logger.Info("Log Name: " + str(self.PeakCan.Logger.fileName))
        self.PeakCan.Logger.Info("Device Name (to be connected): " + str(self.sDevName))
        self.PeakCan.Logger.Info("Bluetooth address(to be connected): " + str(self.iAddress))  # Todo machen
        self.PeakCan.Logger.Info("AutoConnect?: " + str(self.bSthAutoConnect))
        self.PeakCan.Logger.Info("Run Time: " + str(self.iRunTime) + "s")
        self.PeakCan.Logger.Info("Inteval Time: " + str(self.iIntervalTime) + "s")
        self.PeakCan.Logger.Info("Display Time: " + str(self.iDisplayTime) + "ms")
        self.PeakCan.Logger.Info("Adc Prescaler/AcquisitionTime/OversamplingRate/Reference(Samples/s): " + str(self.iPrescaler) + "/" + str(AdcAcquisitionTimeReverse[self.iAquistionTime]) + "/" + str(AdcOverSamplingRateReverse[self.iOversampling]) + "/" + str(self.sAdcRef) + "(" + str(self.samplingRate) + ")")
        self.PeakCan.Logger.Info("Acc Config(XYZ/DataSets): " + str(int(self.bAccX)) + str(int(self.bAccY)) + str(int(self.bAccZ)) + "/" + str(DataSetsReverse[self.tAccDataFormat]))
        self.PeakCan.Logger.Info("Voltage Config(XYZ/DataSets): " + str(int(self.bVoltageX)) + str(int(self.bVoltageY)) + str(int(self.bVoltageZ)) + "/" + str(DataSetsReverse[self.tAccDataFormat]) + ("(X=Battery)"))
        
    def _vRunConsoleStartup(self):
        self._vRunConsoleStartupLoggerPrint()
        if False != self.args_dict['show_config']:
            self._vRunConsoleStartupShow()
        if False != self.args_dict['show_products']:
            self.xmlPrintVersions()
        if False != self.args_dict['show_setups']:
            self.xmlPrintSetups()
        if False != self.args_dict['devNameList']:
            pass  


        
    def clear(self): 
        # for windows 
        if os.name == 'nt': 
            _ = os.system('cls') 
      
        # for mac and linux(here, os.name is 'posix') 
        else: 
            _ = os.system('clear')   
            
    def vRunConsoleAutoConnect(self):
        self.clear()
        if "0x0" != self.iAddress:
            self.PeakCan.BlueToothConnectPollingAddress(MyToolItNetworkNr["STU1"], self.iAddress)
        else:
            self.PeakCan.BlueToothConnectPollingName(MyToolItNetworkNr["STU1"], self.sDevName)
        if False != self.PeakCan.bConnected:
            self.vDataAquisition() 
                               
    def vRunConsole(self):
        self._vRunConsoleStartup()
        self.reset()
        if False != self.bSthAutoConnect:
            self.vRunConsoleAutoConnect()             
        self.close()        

           
if __name__ == "__main__":
    watch = myToolItWatch()
    watch.vParserInit()
    watch.vParserConsoleArgumentsPass()
    watch.vRunConsole()
        
