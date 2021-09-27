import argparse
import array
import copy
import multiprocessing
import os
import socket
import struct
import xml.etree.ElementTree as ET
from time import sleep, time
from datetime import datetime
from pathlib import Path

import openpyxl
from can.interfaces.pcan.basic import PCAN_ERROR_OK, PCAN_ERROR_QOVERRUN
from openpyxl.styles import Font

from mytoolit import __version__
from mytoolit.config import settings
from mytoolit.old.network import Network
from mytoolit.old.MyToolItNetworkNumbers import (MyToolItNetworkName,
                                                 MyToolItNetworkNr)
from mytoolit.old.MyToolItCommands import (
    AdcAcquisitionTime,
    AdcOverSamplingRate,
    AdcReference,
    AtvcFormat,
    byte_list_to_int,
    calcSamplingRate,
    DataSets,
    EepromSpecialConfig,
    iBlueToothMacAddr,
    int_to_mac_address,
    int_to_byte_list,
    rreplace,
    MyToolItBlock,
    MyToolItEeprom,
    MyToolItStreaming,
    payload2Hex,
    Prescaler,
    sArray2String,
    SystemCommandBlueTooth,
    SystemCommandRouting,
)
from mytoolit.old.MyToolItSth import TestConfig
from mytoolit.old.configKeys import ConfigKeys
from mytoolit.old.Plotter import vPlotter, tArray2Binary

Watch = {
    "IntervalDimMinX": 10,  # Minimum interval time in ms
    "DisplayTimeMax": 10,  # Maximum display time of graphical plot in seconds
    "DisplaySampleRateMs": 1000,  # Maximum Display Time in ms
    "DisplayBlockSize": 100,
    "AliveTimeOutMs":
    4000,  # Time Out after receiving no data in acquiring mode
}


class myToolItWatch():

    def __init__(self, *args, **kwargs):
        self.vXmlConfigSet('configKeys.xml')
        self.KeyBoardInterrupt = False
        self.bEepromIgnoreReadErrors = False
        self.bError = False
        self.iMsgLoss = 0
        self.iMsgsTotal = 0
        self.iMsgCounterLast = 0
        self.Can = Network(settings.Logger.icoc.filename,
                           log_directory=settings.Logger.icoc.directory,
                           sender=MyToolItNetworkNr["SPU1"],
                           receiver=MyToolItNetworkNr["STH1"])
        # This method call is currently required to add the timestamp to the
        # log file name
        self.bLogSet(settings.Logger.icoc.filename)
        self.vSave2Xml(False)
        self.vSthAutoConnect(False)
        self.Can.Logger.Info("Start Time: " + self.sDateClock())
        self.bSampleSetupSet(None)
        self.vConfigSet(None, None)
        self.vSheetFileSet(None)
        self.vAccSet(True, False, False, 3)
        self.vVoltageSet(False, False, False, 3)
        self.vDeviceNameSet(TestConfig["DevName"])
        self.vDeviceAddressSet("0")
        self.vAdcConfig(2, 8, 64)
        self.vAdcRefVConfig("VDD")
        self.vDisplayTime(10)
        self.vRunTime(0, 0)
        self.vGraphInit(Watch["DisplaySampleRateMs"],
                        Watch["DisplayBlockSize"])
        self.vStuAddr("")
        self.Can.readThreadStop()
        self.vXmlConfigurationPlotterHost()

    def __exit__(self):
        self.guiProcessStop()
        self.Can.ReadThreadReset()
        if self.Can.bConnected:
            self._BlueToothStatistics()
            ReceiveFailCounter = self._RoutingInformation()
            self._statusWords()
            if ReceiveFailCounter > 0:
                self.bError = True
            self.Can.bBlueToothDisconnect(MyToolItNetworkNr["STU1"])
        if self.Can.bConnected:
            self.Can.readThreadStop()
        self.Can.Logger.Info("End Time Stamp")

        if self.bError:
            self.Can.Logger.Info("!!!!Error!!!!")
            print("!!!!Error!!!!")
        else:
            self.Can.Logger.Info("Fin")
        self.Can.__exit__()
        if self.bError:
            raise
        if self.bSave:
            self.xmlSave()

    def vXmlConfigSet(self, sXmlFileName):
        try:
            self.tXmlConfig.close()
        except:
            pass
        self.sXmlFileName = sXmlFileName
        self.tXmlConfig = ConfigKeys(self.sXmlFileName)

    def vXmlConfigurationPlotterHost(self):
        """Set Matplotlib GUI host and port"""
        self.sPloterSocketHost = settings.gui.host
        self.iPloterSocketPort = settings.gui.port

    def _statusWords(self):
        self.Can.Logger.Info("STH Status Word: {}".format(
            self.Can.node_status(MyToolItNetworkNr["STH1"])))
        self.Can.Logger.Info("STU Status Word: {}".format(
            self.Can.node_status(MyToolItNetworkNr["STU1"])))

        status = self.Can.error_status(MyToolItNetworkNr["STH1"])
        if status.adc_overrun():
            self.bError = True
        self.Can.Logger.Info(f"STH Error Word: {status}")

        self.Can.Logger.Info("STU Error Word: {}".format(
            self.Can.error_status(MyToolItNetworkNr["STU1"])))

    def _BlueToothStatistics(self):
        SendCounter = self.Can.BlueToothCmd(
            MyToolItNetworkNr["STH1"], SystemCommandBlueTooth["SendCounter"])
        self.Can.Logger.Info("BlueTooth Send Counter(STH1): " +
                             str(SendCounter))
        Rssi = self.Can.BlueToothRssi(MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("BlueTooth Rssi(STH1): " + str(Rssi) + "dBm")
        SendCounter = self.Can.BlueToothCmd(
            MyToolItNetworkNr["STU1"], SystemCommandBlueTooth["SendCounter"])
        self.Can.Logger.Info("BlueTooth Send Counter(STU1): " +
                             str(SendCounter))
        ReceiveCounter = self.Can.BlueToothCmd(
            MyToolItNetworkNr["STU1"],
            SystemCommandBlueTooth["ReceiveCounter"])
        self.Can.Logger.Info("BlueTooth Receive Counter(STU1): " +
                             str(ReceiveCounter))
        Rssi = self.Can.BlueToothRssi(MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("BlueTooth Rssi(STU1): " + str(Rssi) + "dBm")

    def _RoutingInformationSthSend(self):
        self.iSthSendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STH1"], SystemCommandRouting["SendCounter"],
            MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STH1 - Send Counter(Port STU1): " +
                             str(self.iSthSendCounter))
        SendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STH1"], SystemCommandRouting["SendFailCounter"],
            MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STH1 - Send Fail Counter(Port STU1): " +
                             str(SendCounter))
        SendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STH1"],
            SystemCommandRouting["SendLowLevelByteCounter"],
            MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STH1 - Send Byte Counter(Port STU1): " +
                             str(SendCounter))

    def _RoutingInformationSthReceive(self):
        ReceiveCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STH1"], SystemCommandRouting["ReceiveCounter"],
            MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STH1 - Receive Counter(Port STU1): " +
                             str(ReceiveCounter))
        ReceiveFailCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STH1"],
            SystemCommandRouting["ReceiveFailCounter"],
            MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STH1 - Receive Fail Counter(Port STU1): " +
                             str(ReceiveFailCounter))
        ReceiveCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STH1"],
            SystemCommandRouting["ReceiveLowLevelByteCounter"],
            MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STH1 - Receive Byte Counter(Port STU1): " +
                             str(ReceiveCounter))
        return ReceiveFailCounter

    def _RoutingInformationSth(self):
        self._RoutingInformationSthSend()
        ReceiveFailCounter = self._RoutingInformationSthReceive()
        return ReceiveFailCounter

    def _RoutingInformationStuPortSpuSend(self):
        self.iStuSendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"], SystemCommandRouting["SendCounter"],
            MyToolItNetworkNr["SPU1"])
        self.Can.Logger.Info("STU1 - Send Counter(Port SPU1): " +
                             str(self.iStuSendCounter))
        SendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"], SystemCommandRouting["SendFailCounter"],
            MyToolItNetworkNr["SPU1"])
        self.Can.Logger.Info("STU1 - Send Fail Counter(Port SPU1): " +
                             str(SendCounter))
        SendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"],
            SystemCommandRouting["SendLowLevelByteCounter"],
            MyToolItNetworkNr["SPU1"])
        self.Can.Logger.Info("STU1 - Send Byte Counter(Port SPU1): " +
                             str(SendCounter))

    def _RoutingInformationStuPortSpuReceive(self):
        ReceiveCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveCounter"],
            MyToolItNetworkNr["SPU1"])
        self.Can.Logger.Info("STU1 - Receive Counter(Port SPU1): " +
                             str(ReceiveCounter))
        ReceiveFailCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"],
            SystemCommandRouting["ReceiveFailCounter"],
            MyToolItNetworkNr["SPU1"])
        self.Can.Logger.Info("STU1 - Receive Fail Counter(Port SPU1): " +
                             str(ReceiveFailCounter))
        ReceiveCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"],
            SystemCommandRouting["ReceiveLowLevelByteCounter"],
            MyToolItNetworkNr["SPU1"])
        self.Can.Logger.Info("STU1 - Receive Byte Counter(Port SPU1): " +
                             str(ReceiveCounter))
        return ReceiveFailCounter

    def _RoutingInformationStuPortSpu(self):
        self._RoutingInformationStuPortSpuSend()
        ReceiveFailCounter = self._RoutingInformationStuPortSpuReceive()
        return ReceiveFailCounter

    def _RoutingInformationStuPortSthSend(self):
        iStuSendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"], SystemCommandRouting["SendCounter"],
            MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("STU1 - Send Counter(Port STH1): " +
                             str(iStuSendCounter))
        SendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"], SystemCommandRouting["SendFailCounter"],
            MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("STU1 - Send Fail Counter(Port STH1): " +
                             str(SendCounter))
        SendCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"],
            SystemCommandRouting["SendLowLevelByteCounter"],
            MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("STU1 - Send Byte Counter(Port STH1): " +
                             str(SendCounter))

    def _RoutingInformationStuPortSthReceive(self):
        ReceiveCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"], SystemCommandRouting["ReceiveCounter"],
            MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("STU1 - Receive Counter(Port STH1): " +
                             str(ReceiveCounter))
        ReceiveFailCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"],
            SystemCommandRouting["ReceiveFailCounter"],
            MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("STU1 - Receive Fail Counter(Port STH1): " +
                             str(ReceiveFailCounter))
        ReceiveCounter = self.Can.RoutingInformationCmd(
            MyToolItNetworkNr["STU1"],
            SystemCommandRouting["ReceiveLowLevelByteCounter"],
            MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("STU1 - Receive Byte Counter(Port STH1): " +
                             str(ReceiveCounter))
        return ReceiveFailCounter

    def _RoutingInformationStuPortSth(self):
        self._RoutingInformationStuPortSthSend()
        ReceiveFailCounter = self._RoutingInformationStuPortSthReceive()
        return ReceiveFailCounter

    def _RoutingInformation(self):
        ReceiveFailCounter = self._RoutingInformationSth()
        ReceiveFailCounter += self._RoutingInformationStuPortSth()
        ReceiveFailCounter += self._RoutingInformationStuPortSpu()
        iSendFail = self.iSthSendCounter - self.iStuSendCounter
        if 0 > iSendFail:
            iSendFail = 0
        iSendFail = iSendFail / self.iSthSendCounter
        iSendFail *= 100
        self.Can.Logger.Info("Send fail approximately: " + str(iSendFail) +
                             "%")
        if 0 < iSendFail:
            print("Send fail approximately: " + str(iSendFail) + "%")
        return ReceiveFailCounter

    """
    version
    """

    def sVersion(self):
        return __version__


# Setter Methods

    def vSave2Xml(self, bSave):
        self.bSave = bSave

    def vConfigSet(self, product, sConfig):
        self.sProduct = None
        self.sConfig = None
        if "STH" == product:
            self.sProduct = "STH"
            self.Can.vSetReceiver(MyToolItNetworkNr["STH1"])
        elif "STU" == product:
            self.sProduct = "STU"
            self.Can.vSetReceiver(MyToolItNetworkNr["STU1"])
        self.sConfig = sConfig

    def bSampleSetupSet(self, sSetup):
        bReturn = False
        self.sSetupConfig = sSetup
        for config in self.tXmlConfig.tree.find('Config'):
            if self.sSetupConfig == config.get('name'):
                bReturn = True
                break
        return bReturn

    def bLogSet(self, sLogLocation):
        bOk = False
        if -1 != sLogLocation.rfind('.'):
            sLogLocation = rreplace(sLogLocation, '.',
                                    "_" + self.sDateClock() + ".")
            self.Can.vLogNameChange(sLogLocation)
            bOk = True
        return bOk

    def vLogCountInc(self):
        fileName = self.Can.Logger.filepath.name[:-24]
        fileName = fileName + "_" + self.sDateClock() + ".txt"
        self.Can.vLogNameCloseInterval(fileName)

    def vSheetFileSet(self, sSheetFile):
        self.sSheetFile = sSheetFile

    def vAccSet(self, bX, bY, bZ, dataSets):
        self.bAccX = bool(bX)
        self.bAccY = bool(bY)
        self.bAccZ = bool(bZ)

        if dataSets in DataSets:
            self.tAccDataFormat = DataSets[dataSets]
        else:
            dataSets = self.Can.dataSetsCan20(bX, bY, bZ)
            self.tAccDataFormat = DataSets[dataSets]

    def vVoltageSet(self, bX, bY, bZ, dataSets):
        self.bVoltageX = bool(bX)
        self.bVoltageY = bool(bY)
        self.bVoltageZ = bool(bZ)

        if dataSets in DataSets:
            self.tVoltageDataFormat = DataSets[dataSets]
        else:
            dataSets = self.Can.dataSetsCan20(bX, bY, bZ)
            self.tVoltageDataFormat = DataSets[dataSets]

    def vSthAutoConnect(self, bSthAutoConnect):
        self.bSthAutoConnect = bool(bSthAutoConnect)

    def vDeviceNameSet(self, sDevName):
        if 8 < len(sDevName):
            sDevName = sDevName[:8]
        self.sDevName = sDevName
        self.iDevNr = None

    def vDeviceAddressSet(self, iAddress):
        """Set bluetooth device address"""
        iAddress = int(iAddress, base=0)
        if 0 < iAddress and iAddress < 2**48 - 1:
            iAddress = hex(iAddress)
            self.iAddress = iAddress
        else:
            self.iAddress = 0

    def vAdcConfig(self, iPrescaler, iAquistionTime, iOversampling):
        if Prescaler["Min"] > iPrescaler:
            iPrescaler = Prescaler["Min"]
        elif Prescaler["Max"] < iPrescaler:
            iPrescaler = Prescaler["Max"]
        try:
            iAcquisitionTime = AdcAcquisitionTime[iAquistionTime]
            iOversampling = AdcOverSamplingRate[iOversampling]
            self.samplingRate = int(
                calcSamplingRate(iPrescaler, iAcquisitionTime, iOversampling) +
                0.5)
            self.iPrescaler = iPrescaler
            self.iAquistionTime = iAcquisitionTime
            self.iOversampling = iOversampling
        except:
            pass

    def vAdcRefVConfig(self, sAdcRef):
        self.sAdcRef = sAdcRef

    def vDisplayTime(self, iDisplayTime):
        """Set the length of the graphical plot in seconds"""
        self.iDisplayTime = int(min(iDisplayTime, Watch["DisplayTimeMax"]))

    def vRunTime(self, runTime, intervalTime):
        self.iIntervalTime = int(intervalTime)
        if self.iIntervalTime <= Watch["IntervalDimMinX"]:
            self.iIntervalTime = 0
        self.iRunTime = int(runTime)

    def vGraphInit(self, sampleInterval=200, blockSize=10):
        """
        sampleInterval in ms
        """
        self.tDataPointTimeStamp = 0  # Time stamp of last data point
        self.iPacketLossTimeStamp = 0
        self.iGraphBlockSize = blockSize
        self.iGraphSampleInterval = sampleInterval
        self.sMsgLoss = "Acceleration(" + str(format(0, '3.3f')) + "%)"
        self.GuiPackage = {"X": [], "Y": [], "Z": []}

    def vStuAddr(self, sStuAddr):
        self.sStuAddr = sStuAddr

    def guiProcessStop(self):
        try:
            self.vGraphSend(["Run", False])
            self.tSocket.close()
            self.guiProcess.terminate()
            self.guiProcess.join()
        except:
            pass

    def vGraphSend(self, data):
        bSend = True
        data = tArray2Binary(data)
        while bSend == True:
            self.tSocket.sendall(data)
            sleep(0.1)
            ack = self.tSocket.recv(2**10)
            self.Can.Logger.Info(
                f"{datetime.now().time()}: Received acknowledgment: {ack}")
            if ack is not None and ack == data:
                bSend = False

    def guiProcessRestart(self):
        self.guiProcessStop()
        if 0 < self.iDisplayTime:
            self.guiProcess = multiprocessing.Process(
                target=vPlotter, args=(self.iPloterSocketPort, ))
            self.guiProcess.start()

            # Wait until socket of GUI application is ready
            connection_established = False
            while not connection_established:
                try:
                    self.tSocket = socket.socket(socket.AF_INET,
                                                 socket.SOCK_STREAM)
                    self.tSocket.connect(
                        (self.sPloterSocketHost, self.iPloterSocketPort))
                    connection_established = True
                except ConnectionError:
                    sleep(0.1)

            self.vGraphSend(["dataBlockSize", self.iGraphBlockSize])
            self.vGraphSend(["sampleInterval", self.iGraphSampleInterval])
            self.vGraphSend(["xDim", self.iDisplayTime])
            self.vGraphPacketLossUpdate(0)
            if False != self.bAccX:
                self.vGraphSend(["lineNameX", "AccX"])
            if False != self.bAccY:
                self.vGraphSend(["lineNameY", "AccY"])
            if False != self.bAccZ:
                self.vGraphSend(["lineNameZ", "AccZ"])
            self.vGraphSend(["Plot", True])

    def vGraphPointNext(self, x, y, z):
        if self.iDisplayTime <= 0:
            return

        if self.guiProcess.is_alive():
            timeStampNow = int(round(time() * 1000))
            elapsed_time_ms = timeStampNow - self.tDataPointTimeStamp
            if (self.iGraphSampleInterval / self.iGraphBlockSize <=
                    elapsed_time_ms):
                self.tDataPointTimeStamp = timeStampNow
                self.GuiPackage["X"].append(x)
                self.GuiPackage["Y"].append(y)
                self.GuiPackage["Z"].append(z)
                if self.iGraphBlockSize <= len(self.GuiPackage["X"]):
                    try:
                        self.tSocket.sendall(
                            tArray2Binary(["data", self.GuiPackage]))
                    except:
                        pass
                    self.GuiPackage = {"X": [], "Y": [], "Z": []}
        else:
            self.aquireEndTime = self.Can.get_elapsed_time()

    def vGraphPacketLossUpdate(self, msgCounter):
        if self.iDisplayTime <= 0:
            return

        self.iMsgCounterLast += 1
        self.iMsgCounterLast %= 256
        if self.iMsgCounterLast != msgCounter:
            iLost = msgCounter - self.iMsgCounterLast
            self.iMsgLoss += iLost
            self.iMsgsTotal += iLost
            if 0 > iLost:
                self.iMsgLoss += 256
                self.iMsgsTotal += 256
            self.iMsgCounterLast = msgCounter
        else:
            self.iMsgsTotal += 1
        iPacketLossTimeStamp = int(round(time() * 1000))
        if 1000 <= (iPacketLossTimeStamp - self.iPacketLossTimeStamp):
            self.iPacketLossTimeStamp = iPacketLossTimeStamp
            sMsgLoss = "Acceleration(" + str(
                format(100 -
                       (100 * self.iMsgLoss / self.iMsgsTotal), '3.3f')) + "%)"
            if sMsgLoss != self.sMsgLoss:
                self.sMsgLoss = sMsgLoss
                self.tSocket.sendall(
                    tArray2Binary(["diagramName", self.sMsgLoss]))
            self.iMsgLoss = 0
            self.iMsgsTotal = 0

    def vVersion(self, major, minor, build):
        if 2 <= major and 1 <= minor:
            self.Major = major
            self.Minor = minor
            self.Build = build

    def sDateClock(self):
        DataClockTimeStamp = datetime.fromtimestamp(
            time()).strftime('%Y-%m-%d_%H-%M-%S')
        return DataClockTimeStamp

    def vParserInit(self):
        self.parser = argparse.ArgumentParser(
            description='Command Line Options')
        self.parser.add_argument(
            '-a',
            '--adc',
            dest='adc_config',
            action='store',
            nargs=3,
            type=int,
            required=False,
            help=
            'Prescaler AcquisitionTime OversamplingRate (3 inputs required in that order e.g. 2 8 64) - Note that acceleration and battery voltage measurements share a single ADC that samples up to 4 channels)'
        )
        self.parser.add_argument(
            '-b',
            '--bluetooth_connect',
            dest='bluetooth_connect',
            action='store',
            nargs=1,
            type=str,
            required=False,
            help=
            'Connect to device specified by Bluetooth address and starts sampling as configured'
        )
        self.parser.add_argument(
            '-d, --gui_dim',
            dest='gui_dim',
            action='store',
            nargs=1,
            type=int,
            required=False,
            help=
            'Length of visualization interval in ms for the graphical acceleration view . Value below 10 turns it off'
        )
        self.parser.add_argument(
            '-e',
            '--xlsx',
            dest='xlsx',
            action='store',
            nargs=1,
            type=str,
            required=False,
            help=
            'Table Calculation File(xlsx) name for transferring data between PC and STH/STU'
        )
        self.parser.add_argument(
            '-i',
            '--interval',
            dest='interval',
            action='store',
            nargs=1,
            type=int,
            required=False,
            help=
            'Sets Interval Time (Output file is saved each interval time in seconds. Lower than 10 causes a single file'
        )
        self.parser.add_argument(
            '-l',
            '--log_name',
            type=str,
            required=False,
            help="(Base) name of the (acceleration) log file")
        self.parser.add_argument(
            '-n',
            '--name_connect',
            dest='name_connect',
            action='store',
            nargs=1,
            type=str,
            required=False,
            help=
            'Connect to device specified by Name and starts sampling as configured'
        )
        self.parser.add_argument(
            '-p',
            '--points',
            dest='points',
            action='store',
            nargs=1,
            type=int,
            required=False,
            help=
            'PPP specifies which acceleration axis(X/Y/Z) are active(1) or off(0)'
        )
        self.parser.add_argument(
            '-r',
            '--run_time',
            dest='run_time',
            action='store',
            nargs=1,
            type=int,
            required=False,
            help='Sets RunTime in seconds. Below 10 specifies infinity')
        self.parser.add_argument(
            '-s',
            '--sample_setup',
            dest='sample_setup',
            action='store',
            nargs=1,
            type=str,
            required=False,
            help=
            'Starts sampling with configuration as given including additional command line arguments'
        )
        self.parser.add_argument(
            '-v',
            '--version',
            dest='version',
            action='store',
            nargs=2,
            type=str,
            required=False,
            help=
            'Chooses product with version for handling Table Calculation Files (e.g. STH v2.1.2)'
        )
        self.parser.add_argument(
            '-x',
            '--xml',
            dest='xml_file_name',
            action='store',
            nargs=1,
            type=str,
            help='Selects xml configuration/data base file',
            default=['configKeys.xml'])
        self.parser.add_argument('--refv',
                                 dest='refv',
                                 action='store',
                                 nargs=1,
                                 type=str,
                                 required=False,
                                 help='ADC\'s Reference voltage, VDD=Standard')
        self.parser.add_argument(
            '--save',
            dest='save',
            action='store_true',
            required=False,
            help='Saves a device configuration or sample setup in the xml file)'
        )
        self.parser.add_argument(
            '--show_config',
            dest='show_config',
            action='store_true',
            required=False,
            help=
            'Shows current configuration (including command line arguments)')
        self.parser.add_argument(
            '--show_products',
            dest='show_products',
            action='store_true',
            required=False,
            help='Shows all available devices and additional versions')
        self.parser.add_argument(
            '--show_setups',
            dest='show_setups',
            action='store_true',
            required=False,
            help=
            'Shows current configuration (including command line arguments)')
        self.parser.add_argument(
            '--voltage_points',
            dest='voltage_points',
            action='store',
            nargs=1,
            type=int,
            required=False,
            help=
            'PPP specifies which voltage axis (sample points; X/Y/Z) are active(1) or off(0). Note that x specifies the battery'
        )
        args = self.parser.parse_args()
        self.args_dict = vars(args)

    def vParserConsoleArgumentsPassXml(self):
        if None != self.args_dict['version'] and None != self.args_dict[
                'sample_setup']:
            print(
                "You can't use sample setup and product/version simultaneously"
            )
            self.Can.vLogDel()
            self.__exit__()
        self.vXmlConfigSet(self.args_dict['xml_file_name'][0])
        if False != self.args_dict['save']:
            self.vSave2Xml(True)
        if None != self.args_dict['sample_setup']:
            sSetup = self.args_dict['sample_setup'][0]
            if False != self.bSampleSetupSet(sSetup):
                self.vGetXmlSetup()
        if None != self.args_dict['version']:
            self.vConfigSet(self.args_dict['version'][0],
                            self.args_dict['version'][1])

    def vParserConsoleArgumentsPass(self):
        self.vParserConsoleArgumentsPassXml()
        if None != self.args_dict['gui_dim']:
            self.vDisplayTime(self.args_dict['gui_dim'][0])
        if self.args_dict['log_name'] is not None:
            self.bLogSet(self.args_dict['log_name'])
        if None != self.args_dict['adc_config']:
            adcConfig = self.args_dict['adc_config']
            self.vAdcConfig(adcConfig[0], adcConfig[1], adcConfig[2])
        if None != self.args_dict['refv']:
            self.vAdcRefVConfig(self.args_dict['refv'][0])
        if None != self.args_dict['xlsx']:
            self.vSheetFileSet(self.args_dict['xlsx'][0])
        iIntervalTime = self.iIntervalTime
        if None != self.args_dict['interval']:
            iIntervalTime = self.args_dict['interval'][0]
        iRunTime = self.iRunTime
        if None != self.args_dict['run_time']:
            iRunTime = self.args_dict['run_time'][0]
        self.vRunTime(iRunTime, iIntervalTime)

        if None != self.args_dict['name_connect']:
            self.vDeviceNameSet(self.args_dict['name_connect'][0])
            self.vSthAutoConnect(True)
        elif None != self.args_dict['bluetooth_connect']:
            int_to_mac_address = str(
                iBlueToothMacAddr(self.args_dict['bluetooth_connect'][0]))
            self.vDeviceAddressSet(int_to_mac_address)
            self.vSthAutoConnect(True)

        if None != self.args_dict['points']:
            points = self.args_dict['points'][0] & 0x07
            bZ = bool(points & 1)
            bY = bool((points >> 1) & 1)
            bX = bool((points >> 2) & 1)
            self.vAccSet(bX, bY, bZ, -1)
        if None != self.args_dict['voltage_points']:
            points = self.args_dict['voltage_points'][0] & 0x07
            bZ = bool(points & 1)
            bY = bool((points >> 1) & 1)
            bX = bool((points >> 2) & 1)
            self.vVoltageSet(bX, bY, bZ, -1)

    def reset(self):
        if False == self.KeyBoardInterrupt:
            try:
                self.Can.ReadThreadReset()
                self.Can.reset_node("STU1")
                self.vStuAddr(
                    int_to_mac_address(
                        self.Can.BlueToothAddress(MyToolItNetworkNr["STU1"])))
                self.guiProcessStop()
            except KeyboardInterrupt:
                self.KeyBoardInterrupt = True

    def vDataAquisition(self):
        if False == self.KeyBoardInterrupt:
            try:
                if self.Can.bConnected:
                    self.Can.ConfigAdc(MyToolItNetworkNr["STH1"],
                                       self.iPrescaler, self.iAquistionTime,
                                       self.iOversampling,
                                       AdcReference[self.sAdcRef])
                    self.Can.readThreadStop()
                    self.guiProcessRestart()
                    self.Can.Logger.Info("Start Acquiring Data")
                    self.vGetStreamingAccData()
                else:
                    self.Can.Logger.Error("Device not allocable")
            except KeyboardInterrupt:
                self.KeyBoardInterrupt = True
            self.__exit__()

    def close(self):
        if False != self.Can.RunReadThread:
            self.__exit__()

    def vGetStreamingAccDataProcess(self):
        iIntervalTime = self.iIntervalTime * 1000
        if 0 == self.iIntervalTime:
            iIntervalTime += (1 << 32)
        startTime = self.Can.get_elapsed_time()
        tAliveTimeStamp = startTime
        tTimeStamp = startTime
        try:
            while (tTimeStamp < self.aquireEndTime):
                try:
                    if (tTimeStamp - startTime) >= iIntervalTime:
                        startTime = tTimeStamp
                        self.vLogCountInc()
                    ack = self.ReadMessage()
                    if (None != ack):
                        tAliveTimeStamp = self.Can.get_elapsed_time()
                        if (self.AccAckExpected.ID != ack["CanMsg"].ID
                                and self.VoltageAckExpected.ID !=
                                ack["CanMsg"].ID):
                            self.Can.Logger.Error("CanId bError: " +
                                                  str(ack["CanMsg"].ID))
                        elif (self.AccAckExpected.DATA[0] !=
                              ack["CanMsg"].DATA[0]
                              and self.VoltageAckExpected.DATA[0] !=
                              ack["CanMsg"].DATA[0]):
                            self.Can.Logger.Error(
                                "Wrong Subheader-Format(Acceleration Format): "
                                + str(ack["CanMsg"].ID))
                        elif self.AccAckExpected.ID == ack["CanMsg"].ID:
                            self.GetMessageAcc(ack)
                        else:
                            self.GetMessageVoltage(ack)
                    else:
                        tTimeStamp = self.Can.get_elapsed_time()
                        if (tAliveTimeStamp +
                                Watch["AliveTimeOutMs"]) < tTimeStamp:
                            self.Can.bConnected = False
                            self.aquireEndTime = tTimeStamp
                            self.Can.Logger.Error(
                                "Not received any streaming package for 4s. Terminated program execution."
                            )
                            print(
                                "Not received any streaming package for 4s. Terminated program execution."
                            )
                except KeyboardInterrupt:
                    pass
            self.__exit__()
        except KeyboardInterrupt:
            self.KeyBoardInterrupt = True
            print("Data acquisition determined")
            self.__exit__()

    def vGetStreamingAccDataAccStart(self):
        ack = None
        if False != self.bAccX or False != self.bAccY or False != self.bAccZ:
            accFormat = AtvcFormat()
            accFormat.asbyte = 0
            accFormat.b.bStreaming = 1
            accFormat.b.bNumber1 = self.bAccX
            accFormat.b.bNumber2 = self.bAccY
            accFormat.b.bNumber3 = self.bAccZ
            accFormat.b.u3DataSets = self.tAccDataFormat
            cmd = self.Can.CanCmd(MyToolItBlock["Streaming"],
                                  MyToolItStreaming["Acceleration"], 0, 0)
            self.AccAckExpected = self.Can.CanMessage20(
                cmd, MyToolItNetworkNr["STH1"], MyToolItNetworkNr["SPU1"],
                [accFormat.asbyte])
            cmd = self.Can.CanCmd(MyToolItBlock["Streaming"],
                                  MyToolItStreaming["Acceleration"], 1, 0)
            message = self.Can.CanMessage20(cmd, MyToolItNetworkNr["SPU1"],
                                            MyToolItNetworkNr["STH1"],
                                            [accFormat.asbyte])
            self.Can.Logger.Info("MsgId/Subpayload(Acc): " + hex(message.ID) +
                                 "/" + hex(accFormat.asbyte))
            endTime = self.Can.get_elapsed_time() + 4000
            while (None == ack) and (self.Can.get_elapsed_time() < endTime):
                self.Can.WriteFrame(message)
                readEndTime = self.Can.get_elapsed_time() + 500
                while ((None == ack)
                       and (self.Can.get_elapsed_time() < readEndTime)):
                    ack = self.ReadMessage()
        else:
            ack = True
        return ack

    def vGetStreamingAccDataVoltageStart(self):
        ack = None
        if False != self.bVoltageX or False != self.bVoltageY or False != self.bVoltageZ:
            voltageFormat = AtvcFormat()
            voltageFormat.asbyte = 0
            voltageFormat.b.bStreaming = 1
            voltageFormat.b.bNumber1 = self.bVoltageX
            voltageFormat.b.bNumber2 = self.bVoltageY
            voltageFormat.b.bNumber3 = self.bVoltageZ
            voltageFormat.b.u3DataSets = self.tVoltageDataFormat
            cmd = self.Can.CanCmd(MyToolItBlock["Streaming"],
                                  MyToolItStreaming["Voltage"], 0, 0)
            self.VoltageAckExpected = self.Can.CanMessage20(
                cmd, MyToolItNetworkNr["STH1"], MyToolItNetworkNr["SPU1"],
                [voltageFormat.asbyte])
            cmd = self.Can.CanCmd(MyToolItBlock["Streaming"],
                                  MyToolItStreaming["Voltage"], 1, 0)
            message = self.Can.CanMessage20(cmd, MyToolItNetworkNr["SPU1"],
                                            MyToolItNetworkNr["STH1"],
                                            [voltageFormat.asbyte])
            self.Can.Logger.Info("MsgId/Subpayload(Voltage): " +
                                 hex(message.ID) + "/" +
                                 hex(voltageFormat.asbyte))

            endTime = self.Can.get_elapsed_time() + 4000
            while (None == ack) and (self.Can.get_elapsed_time() < endTime):
                self.Can.WriteFrame(message)
                readEndTime = self.Can.get_elapsed_time() + 500
                while ((None == ack)
                       and (self.Can.get_elapsed_time() < readEndTime)):
                    ack = self.ReadMessage()
        else:
            ack = True
        return ack

    def vGetStreamingAccData(self):
        ack = self.vGetStreamingAccDataAccStart()
        if None != ack:
            ack = self.vGetStreamingAccDataVoltageStart()
        currentTime = self.Can.get_elapsed_time()
        if None == ack:
            self.Can.Logger.Error("No Ack received from Device: " +
                                  str(self.iDevNr))
            self.aquireEndTime = currentTime
        elif (0 == self.iRunTime):
            self.aquireEndTime = currentTime + (1 << 32)
        else:
            self.aquireEndTime = currentTime + self.iRunTime * 1000
        self.vGetStreamingAccDataProcess()

    def GetMessageSingle(self, prefix, canMsg):
        canData = canMsg["CanMsg"].DATA
        p1 = byte_list_to_int(canData[2:4])
        p2 = byte_list_to_int(canData[6:8])
        p3 = byte_list_to_int(canData[4:6])

        canTimeStamp = canMsg["PeakCanTime"]
        canTimeStamp = round(canTimeStamp, 3)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix + ": ")
        ackMsg += str(format(p1, '5d'))
        ackMsg += "; "
        self.Can.Logger.Info(ackMsg)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix + ": ")
        ackMsg += str(format(p2, '5d'))
        ackMsg += "; "
        self.Can.Logger.Info(ackMsg)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix + ": ")
        ackMsg += str(format(p3, '5d'))
        ackMsg += "; "
        self.Can.Logger.Info(ackMsg)

    def GetMessageDouble(self, prefix1, prefix2, canMsg):
        canData = canMsg["CanMsg"].DATA
        canTimeStamp = canMsg["PeakCanTime"]
        canTimeStamp = round(canTimeStamp, 3)
        p1_1 = byte_list_to_int(canData[2:4])
        p1_2 = byte_list_to_int(canData[4:6])
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix1 + ": ")
        ackMsg += str(format(p1_1, '5d'))
        ackMsg += "; "
        ackMsg += prefix2
        ackMsg += ": "
        ackMsg += str(format(p1_2, '5d'))
        ackMsg += "; "
        self.Can.Logger.Info(ackMsg)

    def GetMessageTripple(self, prefix1, prefix2, prefix3, canMsg):
        canData = canMsg["CanMsg"].DATA
        canTimeStamp = canMsg["PeakCanTime"]
        canTimeStamp = round(canTimeStamp, 3)
        ackMsg = ("MsgCounter: " + str(format(canData[1], '3d')) + "; ")
        ackMsg += ("TimeStamp: " + format(canTimeStamp, '12.3f') + "ms; ")
        ackMsg += (prefix1 + ": ")
        ackMsg += str(format(byte_list_to_int(canData[2:4]), '5d'))
        ackMsg += "; "
        ackMsg += prefix2
        ackMsg += ": "
        ackMsg += str(format(byte_list_to_int(canData[4:6]), '5d'))
        ackMsg += "; "
        ackMsg += prefix3
        ackMsg += ": "
        ackMsg += str(format(byte_list_to_int(canData[6:8]), '5d'))
        ackMsg += "; "
        self.Can.Logger.Info(ackMsg)

    def GetMessageAcc(self, canData):
        data = canData["CanMsg"].DATA
        msgCounter = data[1]
        self.vGraphPacketLossUpdate(msgCounter)

        if self.tAccDataFormat == DataSets[1]:
            if (False != self.bAccX) and (False != self.bAccY) and (
                    False == self.bAccZ):
                self.GetMessageDouble("AccX", "AccY", canData)
                self.vGraphPointNext(byte_list_to_int(data[2:4]),
                                     byte_list_to_int(data[4:6]), 0)
            elif (False != self.bAccX) and (False == self.bAccY) and (
                    False != self.bAccZ):
                self.GetMessageDouble("AccX", "AccZ", canData)
                self.vGraphPointNext(byte_list_to_int(data[2:4]), 0,
                                     byte_list_to_int(data[4:6]))
            elif (False == self.bAccX) and (False != self.bAccY) and (
                    False != self.bAccZ):
                self.GetMessageDouble("AccY", "AccZ", canData)
                self.vGraphPointNext(0, byte_list_to_int(data[2:4]),
                                     byte_list_to_int(data[4:6]))
            else:
                self.GetMessageTripple("AccX", "AccY", "AccZ", canData)
                self.vGraphPointNext(byte_list_to_int(data[2:4]),
                                     byte_list_to_int(data[4:6]),
                                     byte_list_to_int(data[6:8]))
        elif self.tAccDataFormat == DataSets[3]:
            if False != self.bAccX:
                self.GetMessageSingle("AccX", canData)
                self.vGraphPointNext(byte_list_to_int(data[2:4]), 0, 0)
            elif False != self.bAccY:
                self.GetMessageSingle("AccY", canData)
                self.vGraphPointNext(0, byte_list_to_int(data[2:4]), 0)
            elif False != self.bAccZ:
                self.GetMessageSingle("AccZ", canData)
                self.vGraphPointNext(0, 0, byte_list_to_int(data[2:4]))
        else:
            self.Can.Logger.Error("Wrong Ack format")

    def GetMessageVoltage(self, canData):
        if self.tAccDataFormat == DataSets[1]:
            if (0 != self.bVoltageX) and (0 != self.bVoltageY) and (
                    0 == self.bVoltageZ):
                self.GetMessageDouble("VoltageX", "VoltageY", canData)
            elif (0 != self.bVoltageX) and (0 == self.bVoltageY) and (
                    0 != self.bVoltageZ):
                self.GetMessageDouble("VoltageX", "VoltageZ", canData)
            elif (0 == self.bVoltageX) and (0 != self.bVoltageY) and (
                    0 != self.bVoltageZ):
                self.GetMessageDouble("VoltageY", "VoltageZ", canData)
            else:
                self.GetMessageTripple("VoltageX", "VoltageY", "VoltageZ",
                                       canData)
        elif self.tVoltageDataFormat == DataSets[3]:
            if 0 != self.bVoltageX:
                self.GetMessageSingle("VoltageX", canData)
            elif 0 != self.bVoltageY:
                self.GetMessageSingle("VoltageY", canData)
            elif 0 != self.bVoltageZ:
                self.GetMessageSingle("VoltageZ", canData)
        else:
            self.Can.Logger.Error("Wrong Ack format")

    def ReadMessage(self):
        message = None
        result = self.Can.pcan.Read(self.Can.m_PcanHandle)
        if result[0] == PCAN_ERROR_OK:
            peakCanTimeStamp = result[2].millis_overflow * (
                2**32) + result[2].millis + result[2].micros / 1000
            message = {
                "CanMsg": result[1],
                "PcTime": self.Can.get_elapsed_time(),
                "PeakCanTime": peakCanTimeStamp
            }
        elif result[0] == PCAN_ERROR_QOVERRUN:
            self.Can.Logger.Error("RxOverRun")
            print("RxOverRun")
            raise Exception("RxOverRun")
        else:
            sleep(0.0002)
        return message

    def atXmlProductVersion(self):
        dataDef = self.tXmlConfig.root.find('Data')
        atProducts = {}
        iProduct = 1
        for product in dataDef.find('Product'):
            atProducts[iProduct] = {}
            atProducts[iProduct]["Product"] = product
            atProducts[iProduct]["Versions"] = {}
            iVersion = 1
            for version in product.find('Version'):
                atProducts[iProduct]["Versions"][iVersion] = version
                iVersion += 1
            iProduct += 1
        return atProducts

    def excelCellWidthAdjust(self, worksheet, factor=1.2, bSmaller=True):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                    continue
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * factor
            columLetter = chr(ord('A') + column - 1)
            if adjusted_width > worksheet.column_dimensions[
                    columLetter].width or False == bSmaller:
                worksheet.column_dimensions[columLetter].width = adjusted_width

    """
    Get Page Names from xml by product and versions as List - Version
    """

    def atProductPagesVersion(self, version):
        atPageList = []
        if version.get('name') == self.sConfig:
            for page in version.find('Page'):
                tPageDict = {}
                tPageDict["Name"] = str(page.get('name'))
                tPageDict["Address"] = int(page.find('pageAddress').text)
                tPageDict["Entry"] = page.find('Entry')
                atPageList.append(tPageDict)
        return atPageList

    """
    Get Page Names from xml by product and versions as List - Product
    """

    def atProductPagesProduct(self, product):
        atPageList = []
        if None != self.sConfig:
            for version in product.find('Version'):
                if version.get('name') == self.sConfig:
                    atPageList = self.atProductPagesVersion(version)
                    break
        return atPageList

    """
    Get Page Names from xml by product and versions as List
    """

    def atProductPages(self):
        atPageList = []
        if None != self.sProduct:
            dataDef = self.tXmlConfig.root.find('Data')
            for product in dataDef.find('Product'):
                if product.get('name') == self.sProduct:
                    atPageList = self.atProductPagesProduct(product)
                    break
        return atPageList

    def tExcelWorkSheetCreate(self, workbook, name, pageAddress):
        tWorkSheet = workbook.create_sheet(name + "@" + hex(pageAddress))
        tFontRow = Font(bold=True, size=20)
        tWorkSheet['A1'] = 'Name'
        tWorkSheet['A1'].font = tFontRow
        tWorkSheet['B1'] = 'Address'
        tWorkSheet['B1'].font = tFontRow
        tWorkSheet['C1'] = 'Length'
        tWorkSheet['C1'].font = tFontRow
        tWorkSheet['D1'] = 'Read Only'
        tWorkSheet['D1'].font = tFontRow
        tWorkSheet['E1'] = 'Value'
        tWorkSheet['E1'].font = tFontRow
        tWorkSheet['F1'] = 'Unit'
        tWorkSheet['F1'].font = tFontRow
        tWorkSheet['G1'] = 'Format'
        tWorkSheet['G1'].font = tFontRow
        tWorkSheet['H1'] = 'Description'
        tWorkSheet['H1'].font = tFontRow
        return tWorkSheet

    """
    Create Excel Sheet by xml definition
    """

    def vExcelSheetCreate(self):
        atProductPages = self.atProductPages()
        if 0 < len(atProductPages):
            workbook = openpyxl.Workbook()
            tFontRowRow = Font(bold=False, size=12)
            workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
            for page in atProductPages:
                i = 2
                name = page["Name"]
                pageAddress = page["Address"]
                tWorkSheet = self.tExcelWorkSheetCreate(
                    workbook, name, pageAddress)
                self.excelCellWidthAdjust(tWorkSheet, 1.6, False)
                for entry in page["Entry"]:
                    tWorkSheet['A' + str(i)] = entry.get('name')
                    tWorkSheet['A' + str(i)].font = tFontRowRow
                    tWorkSheet['B' + str(i)] = int(
                        entry.find('subAddress').text)
                    tWorkSheet['B' + str(i)].font = tFontRowRow
                    tWorkSheet['C' + str(i)] = int(entry.find('length').text)
                    tWorkSheet['C' + str(i)].font = tFontRowRow
                    tWorkSheet['D' + str(i)] = entry.find('readOnly').text
                    tWorkSheet['D' + str(i)].font = tFontRowRow
                    try:
                        tWorkSheet['E' + str(i)] = int(
                            entry.find('value').text)
                    except ValueError:
                        tWorkSheet['E' + str(i)] = entry.find('value').text
                    tWorkSheet['E' + str(i)].font = tFontRowRow
                    tWorkSheet['F' + str(i)] = entry.find('unit').text
                    tWorkSheet['F' + str(i)].font = tFontRowRow
                    tWorkSheet['G' + str(i)] = entry.find('format').text
                    tWorkSheet['G' + str(i)].font = tFontRowRow
                    tWorkSheet['H' + str(i)] = entry.find('description').text
                    tWorkSheet['H' + str(i)].font = tFontRowRow
                    i += 1
                self.excelCellWidthAdjust(tWorkSheet)
            workbook.save(self.sSheetFile)

    def _excelSheetEntryFind(self, entry, key, value):
        if None != value:
            entry.find(key).text = str(value)

    """
    Set encoding
    """

    def _XmlWriteEndoding(self):
        xml = (bytes('<?xml version="1.0" encoding="UTF-8"?>\n',
                     encoding='utf-8') + ET.tostring(self.tXmlConfig.root))
        xml = xml.decode('utf-8')
        filepath = Path(__file__).parent.joinpath(self.sXmlFileName)
        with open(filepath, "w", encoding='utf-8') as f:
            f.write(xml)

    """
    Creates a new config
    """

    def newXmlVersion(self, product, productVersion, sVersion):
        cloneVersion = copy.deepcopy(productVersion)
        cloneVersion.set('name', sVersion)
        product.find('Version').append(cloneVersion)
        self.xmlSave()
        self.vConfigSet(product.get('name'), sVersion)

    """
    Save XML File (in any state)
    """

    def xmlSave(self):
        filepath = str(Path(__file__).parent.joinpath(self.sXmlFileName))
        self.tXmlConfig.tree.write(filepath)
        self._XmlWriteEndoding()
        del self.tXmlConfig
        self.tXmlConfig = ConfigKeys(self.args_dict['xml_file_name'][0])

    """
    Removes a config
    """

    def removeXmlVersion(self, product, vesion):
        product.find('Version').remove(vesion)
        self.xmlSave()

    def xmlPrintVersions(self):
        dataDef = self.tXmlConfig.root.find('Data')
        for product in dataDef.find('Product'):
            print(product.get('name') + ":")
            for version in product.find('Version'):
                print("   " + version.get('name'))

    """
    Write xml entry by Excel Sheet entry
    """

    def _vExcelProductVersion2XmlProductVersionEntry(self, tEntryXml,
                                                     tWorkSheet, iEntryExcel):
        sExcelEntryName = str(tWorkSheet['A' + str(iEntryExcel)].value)
        if None != sExcelEntryName:
            tEntryXml.set('name', sExcelEntryName)
        else:
            tEntryXml.set('name', "")
        self._excelSheetEntryFind(tEntryXml, 'subAddress',
                                  tWorkSheet['B' + str(iEntryExcel)].value)
        self._excelSheetEntryFind(tEntryXml, 'length',
                                  tWorkSheet['C' + str(iEntryExcel)].value)
        self._excelSheetEntryFind(tEntryXml, 'readOnly',
                                  tWorkSheet['D' + str(iEntryExcel)].value)
        self._excelSheetEntryFind(tEntryXml, 'value',
                                  tWorkSheet['E' + str(iEntryExcel)].value)
        self._excelSheetEntryFind(tEntryXml, 'unit',
                                  tWorkSheet['F' + str(iEntryExcel)].value)
        self._excelSheetEntryFind(tEntryXml, 'format',
                                  tWorkSheet['G' + str(iEntryExcel)].value)
        self._excelSheetEntryFind(tEntryXml, 'description',
                                  tWorkSheet['H' + str(iEntryExcel)].value)

    """
    Create xml entry by Excel Sheet entry
    """

    def _tExcelProductVersion2XmlProductVersionEntryNew(self, atEntries):
        tEntry = self.tXmlChildNew(atEntries, 'entry')
        self.tXmlChildNew(tEntry, 'subAddress')
        self.tXmlChildNew(tEntry, 'length')
        self.tXmlChildNew(tEntry, 'readOnly')
        self.tXmlChildNew(tEntry, 'value')
        self.tXmlChildNew(tEntry, 'unit')
        self.tXmlChildNew(tEntry, 'format')
        self.tXmlChildNew(tEntry, 'description')
        return tEntry

    """
    Write xml definition by Excel Sheet - Excel Page Entries
    """

    def _vExcelProductVersion2XmlProductVersionPageEntries(
            self, tWorkSheet, atXmlEntries):
        iEntryChild = 0
        iExcelRow = 2
        while None != tWorkSheet['A' + str(iExcelRow)].value:
            if iEntryChild < len(atXmlEntries):
                tEntry = atXmlEntries[iEntryChild]
            else:
                tEntry = self._tExcelProductVersion2XmlProductVersionEntryNew(
                    atXmlEntries)
            self._vExcelProductVersion2XmlProductVersionEntry(
                tEntry, tWorkSheet, iExcelRow)
            iEntryChild += 1
            iExcelRow += 1
        # Remove Delete entries
        while iEntryChild < len(atXmlEntries):
            atXmlEntries.remove(atXmlEntries[iEntryChild])
            iEntryChild += 1

    """
    Write existing xml definition by Excel Sheet - Excel Entries
    """

    def _bExcelProductVersion2XmlProductVersionPageExist(
            self, tWorkSheet, atProductPages, sName, sAddress):
        bFound = False
        for tPageDict in atProductPages:
            sXmlName = tPageDict["Name"]
            sXmlAddress = hex(tPageDict["Address"])
            if sName == sXmlName and sXmlAddress == sAddress:
                self._vExcelProductVersion2XmlProductVersionPageEntries(
                    tWorkSheet, tPageDict["Entry"])
                bFound = True
                break
        return bFound

    """
    Create xml page by Excel Work Sheet
    """

    def _vExcelProductVersion2XmlProductVersionPageNew(self, sName, sAddress):
        if None != self.sProduct and None != self.sConfig:
            dataDef = self.tXmlConfig.root.find('Data')
            for product in dataDef.find('Product'):
                if product.get('name') == self.sProduct:
                    for version in product.find('Version'):
                        if version.get('name') == self.sConfig:
                            tPage = self.tXmlChildNew(version.find('Page'),
                                                      'page')
                            tPage.set('name', sName)
                            self.tXmlChildNew(tPage, 'pageAddress')
                            pageAddress = str(int(sAddress, 16))
                            tPage.find('pageAddress').text = pageAddress
                            self.tXmlChildNew(tPage, 'Entry')
                            ET.dump(version.find('Page'))
                            self.xmlSave()
                            break

    """
    Write xml definition by Excel Sheet
    """

    def vExcelProductVersion2XmlProductVersionPage(self, tWorkbook):
        for tWorksheetName in tWorkbook.sheetnames:
            sName = str(tWorksheetName).split('@')
            sAddress = sName[1]
            sName = sName[0]
            tWorkSheet = tWorkbook.get_sheet_by_name(tWorksheetName)
            try:
                atProductPages = self.atProductPages()
            except:
                self._vExcelProductVersion2XmlProductVersionPageNew(
                    sName, sAddress)
                atProductPages = self.atProductPages()
            if False == self._bExcelProductVersion2XmlProductVersionPageExist(
                    tWorkSheet, atProductPages, sName, sAddress):
                self._vExcelProductVersion2XmlProductVersionPageNew(
                    sName, sAddress)
                atProductPages = self.atProductPages()
                if False == self._bExcelProductVersion2XmlProductVersionPageExist(
                        tWorkSheet, atProductPages, sName, sAddress):
                    break

    """
    Create xml page by Excel Work Sheet
    """

    def _vExcelProductVersion2XmlProductVersionXmlPageRemoveAction(
            self, sName):
        if None != self.sProduct and None != self.sConfig:
            dataDef = self.tXmlConfig.root.find('Data')
            for product in dataDef.find('Product'):
                if product.get('name') == self.sProduct:
                    for version in product.find('Version'):
                        if version.get('name') == self.sConfig:
                            for page in version.find('Page'):
                                if page.get('name') == sName:
                                    version.find('Page').remove(page)

    """
    Write xml definition by Excel Sheet - Remove entries that are not part of an excel sheet
    """

    def _vExcelProductVersion2XmlProductVersionXmlPageRemove(self, tWorkbook):
        atProductPages = self.atProductPages()  # Reload to have up2Date Copy
        for i in range(0, len(atProductPages)):
            tPageDict = atProductPages[i]
            sXmlName = tPageDict["Name"]
            sXmlAddress = hex(tPageDict["Address"])
            bFound = False
            for tWorksheetName in tWorkbook.sheetnames:
                sName = str(tWorksheetName).split('@')
                sAddress = sName[1]
                sName = sName[0]
                if sName == sXmlName and sXmlAddress == sAddress:
                    bFound = True
                    break
            if False == bFound:
                self._vExcelProductVersion2XmlProductVersionXmlPageRemoveAction(
                    sXmlName)

    """
    Write xml definition by Excel Sheet and do checks
    """

    def vExcelProductVersion2XmlProductVersion(self):
        if None != self.sProduct and None != self.sConfig and None != self.sSheetFile:
            tWorkbook = openpyxl.load_workbook(self.sSheetFile)
            if tWorkbook:
                try:
                    uLength = len(self.atProductPages())
                except:
                    self.vExcelProductVersion2XmlProductVersionPage(tWorkbook)
                    uLength = len(self.atProductPages())
                if 0 < uLength:
                    self.vExcelProductVersion2XmlProductVersionPage(tWorkbook)
                else:
                    for tProduct in self.tXmlConfig.root.find('Data').find(
                            'Product'):
                        if tProduct.get('name') == self.sProduct:
                            for tVersion in tProduct.find('Version'):
                                if tVersion.get('name') == self.sConfig:
                                    self.newXmlVersion(tProduct, tVersion,
                                                       self.sConfig)
                                    self.xmlSave()
                                    self.vExcelProductVersion2XmlProductVersion(
                                    )
                # Remove Deleted Pages
                self._vExcelProductVersion2XmlProductVersionXmlPageRemove(
                    tWorkbook)
                self.xmlSave()

    def iExcelSheetPageLength(self, worksheet):
        totalLength = 0
        for i in range(2, 256 + 2):
            if None != worksheet['A' + str(i)].value:
                length = int(worksheet['C' + str(i)].value)
                totalLength += length
            else:
                break
        return totalLength

    def vUnicodeIllegalRemove(self, value, character):
        while (True):
            try:
                value.remove(character)
            except:
                break
        return value

    def iExcelSheetPageValue(self, worksheet, aBytes):
        i = 2
        iTotalLength = 0
        while len(aBytes) > iTotalLength:
            if None != worksheet['A' + str(i)].value:
                iLength = worksheet['C' + str(i)].value
                value = aBytes[iTotalLength:iTotalLength + iLength]
                if "UTF8" == worksheet['G' + str(i)].value:
                    try:
                        value = self.vUnicodeIllegalRemove(value, 0)
                        value = self.vUnicodeIllegalRemove(value, 255)
                        value = array.array('b', value).tostring().decode(
                            'utf-8', 'replace')
                    except Exception as e:
                        self.Can.Logger.Info(str(e))
                        value = ""
                elif "ASCII" == worksheet['G' + str(i)].value:
                    value = sArray2String(value)
                elif "unsigned" == worksheet['G' + str(i)].value:
                    value = str(byte_list_to_int(value))
                elif "float" == worksheet['G' + str(i)].value:
                    if None != value:
                        pass
                        # value = au8ChangeEndianOrder(value)
                    else:
                        value = 0.0
                    self.Can.Logger.Info("Value from EEPROM: " + str(value))
                    value = bytearray(value)
                    value = struct.unpack('f', value)[0]
                    value = str(value)
                    self.Can.Logger.Info("Value as float: " + str(value))
                else:
                    value = payload2Hex(value)
                value = str(value)
                try:
                    worksheet['E' + str(i)] = str(value)
                except:
                    pass
                iTotalLength += iLength
                i += 1
            else:
                break
        return iTotalLength

    def atExcelSheetNames(self):
        workSheetNames = []
        try:
            workbook = openpyxl.load_workbook(self.sSheetFile)
            if workbook:
                for worksheetName in workbook.sheetnames:
                    sName = str(worksheetName).split('@')
                    sName = sName[0]
                    workSheetNames.append(sName)
        except:
            pass
        return workSheetNames

    """
    Read EEPROM page to write values in Excel Sheet
    """

    def sExcelSheetRead(self, namePage, iReceiver):
        tEepromSpecialConfig = EepromSpecialConfig()
        tEepromSpecialConfig.asbyte = 0
        tEepromSpecialConfig.b.bIgnoreErrors = self.bEepromIgnoreReadErrors
        sError = None
        self.Can.Logger.Info("Read EEPROM Page " + str(namePage) + " from " +
                             MyToolItNetworkName[iReceiver])
        workbook = openpyxl.load_workbook(self.sSheetFile)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = int(name[1], base=16)
                name = name[0]
                if namePage == name:
                    worksheet = workbook.get_sheet_by_name(worksheetName)
                    pageContent = []
                    readLength = self.iExcelSheetPageLength(worksheet)
                    readLengthAlligned = readLength
                    if 0 != readLengthAlligned % 4:
                        readLengthAlligned += 4
                        readLengthAlligned -= (readLengthAlligned % 4)
                    for offset in range(0, readLengthAlligned, 4):
                        payload = [
                            address, 0xFF & offset, 4,
                            tEepromSpecialConfig.asbyte, 0, 0, 0, 0
                        ]
                        index = self.Can.cmdSend(iReceiver,
                                                 MyToolItBlock["EEPROM"],
                                                 MyToolItEeprom["Read"],
                                                 payload,
                                                 log=False)
                        readBackFrame = self.Can.getReadMessageData(index)[4:]
                        pageContent.extend(readBackFrame)
                    pageContent = pageContent[0:readLength]
                    self.Can.Logger.Info("Read Data: " +
                                         payload2Hex(pageContent))
                    self.iExcelSheetPageValue(worksheet, pageContent)
            try:
                workbook.save(self.sSheetFile)
            except Exception as e:
                sError = "Could not save file(Opened by another application?): " + str(
                    e)
                self.Can.Logger.Info(sError)
                print(sError)
        return sError

    def au8excelValueToByteArray(self, worksheet, iIndex):
        iLength = int(worksheet['C' + str(iIndex)].value)
        value = worksheet['E' + str(iIndex)].value
        byteArray = [0] * iLength
        if None != value:
            if "UTF8" == worksheet['G' + str(iIndex)].value:
                try:
                    value = str(value).encode('utf-8')
                except Exception as e:
                    self.Can.Logger.Info(str(e))
                    value = [0] * iLength
                iCopyLength = len(value)
                if iLength < iCopyLength:
                    iCopyLength = iLength
                for i in range(0, iCopyLength):
                    byteArray[i] = value[i]
            elif "ASCII" == worksheet['G' + str(iIndex)].value:
                try:
                    value = str(value).encode('ascii')
                except Exception as e:
                    self.Can.Logger.Info(str(e))
                    value = [0] * iLength
                iCopyLength = len(value)
                if iLength < iCopyLength:
                    iCopyLength = iLength
                for i in range(0, iCopyLength):
                    byteArray[i] = value[i]
            elif "unsigned" == worksheet['G' + str(iIndex)].value:
                byteArray = int_to_byte_list(int(value), iLength)
            elif "float" == worksheet['G' + str(iIndex)].value:
                value = float(value)
                value = struct.pack('f', value)
                value = int.from_bytes(value, byteorder='little')
                byteArray = int_to_byte_list(int(value), 4)
            else:
                if "0" == value or 0 == value:
                    value = "[0x0]"
                value = value[1:-1].split(',')
                for i in range(0, len(value)):
                    byteArray[i] = int(value[i], 16)
        return byteArray

    """
    Read Excel Sheet to write values to EEPROM
    """

    def sExcelSheetWrite(self, namePage, iReceiver):
        sError = None
        self.Can.Logger.Info("Write Excel Page " + str(namePage) + " to " +
                             MyToolItNetworkName[iReceiver])
        workbook = openpyxl.load_workbook(self.sSheetFile)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = int(name[1], base=16)
                name = name[0]
                if namePage == name:
                    worksheet = workbook.get_sheet_by_name(worksheetName)
                    # Prepare Write Data
                    au8WriteData = [0] * 256
                    iByteIndex = 0
                    for i in range(2, 256 + 2, 1):
                        if None != worksheet['A' + str(i)].value:
                            au8ElementData = self.au8excelValueToByteArray(
                                worksheet, i)
                            for j in range(0, len(au8ElementData), 1):
                                au8WriteData[iByteIndex +
                                             j] = au8ElementData[j]
                            iLength = int(worksheet['C' + str(i)].value)
                            iByteIndex += iLength
                        else:
                            break
                    iWriteLength = self.iExcelSheetPageLength(worksheet)
                    if 0 != iWriteLength % 4:
                        iWriteLength += 4
                        iWriteLength -= (iWriteLength % 4)
                    au8WriteData = au8WriteData[0:iWriteLength]
                    self.Can.Logger.Info("Write Content: " +
                                         payload2Hex(au8WriteData))
                    for offset in range(0, iWriteLength, 4):
                        au8WritePackage = au8WriteData[offset:offset + 4]
                        au8Payload = [address, 0xFF & offset, 4, 0]
                        au8Payload.extend(au8WritePackage)
                        self.Can.cmdSend(iReceiver,
                                         MyToolItBlock["EEPROM"],
                                         MyToolItEeprom["Write"],
                                         au8Payload,
                                         log=False)
            try:
                workbook.close()
            except Exception as e:
                sError = "Could not close file: " + str(e)
                self.Can.Logger.Info(sError)
                print(sError)
        return sError

    def atXmlSetup(self):
        asSetups = {}
        iSetup = 1
        for setup in self.tXmlConfig.tree.find('Config'):
            asSetups[iSetup] = setup
            iSetup += 1
        return asSetups

    def vSetXmlSetup(self):
        for config in self.tXmlConfig.tree.find('Config'):
            if config.get('name') == self.sSetupConfig:
                config.find('DeviceName').text = str(self.sDevName)
                config.find('DeviceAddress').text = str(self.iAddress)
                config.find('Acc').text = str(int(self.bAccX)) + str(
                    int(self.bAccY)) + str(int(self.bAccZ))
                config.find('Voltage').text = str(int(self.bVoltageX)) + str(
                    int(self.bVoltageY)) + str(int(self.bVoltageZ))
                config.find('Prescaler').text = str(self.iPrescaler)
                config.find('AcquisitionTime').text = str(
                    AdcAcquisitionTime.inverse[self.iAquistionTime])
                config.find('OverSamples').text = str(
                    AdcOverSamplingRate.inverse[self.iOversampling])
                config.find('AdcRef').text = str(self.sAdcRef)
                sFileName = self.Can.Logger.filepath.name
                config.find('LogName').text = sFileName[:sFileName.find('_'):]
                config.find('RunTime').text = str(self.iRunTime)
                config.find('IntervalTime').text = str(self.iIntervalTime)
                config.find('DisplayTime').text = str(self.iDisplayTime)
                break

    def vGetXmlSetup(self):
        for config in self.tXmlConfig.tree.find('Config'):
            if config.get('name') == self.sSetupConfig:
                self.vDeviceNameSet(config.find('DeviceName').text)
                self.vDeviceAddressSet(config.find('DeviceAddress').text)
                if ("" != self.sDevName or 0 < self.iAddress):
                    self.vSthAutoConnect(True)
                samplePoints = config.find('Acc').text
                bAccX = int(samplePoints[0])
                bAccY = int(samplePoints[1])
                bAccZ = int(samplePoints[2])
                self.vAccSet(bAccX, bAccY, bAccZ, -1)
                samplePoints = config.find('Voltage').text
                bVoltageX = int(samplePoints[0])
                bVoltageY = int(samplePoints[1])
                bVoltageZ = int(samplePoints[2])
                self.vVoltageSet(bVoltageX, bVoltageY, bVoltageZ, -1)
                self.vAdcConfig(int(config.find('Prescaler').text),
                                int(config.find('AcquisitionTime').text),
                                int(config.find('OverSamples').text))
                self.vAdcRefVConfig(config.find('AdcRef').text)
                self.bLogSet(str(config.find('LogName').text) + ".txt")
                self.vRunTime(int(config.find('RunTime').text),
                              int(config.find('IntervalTime').text))
                self.vDisplayTime(int(config.find('DisplayTime').text))
                break

    def removeXmlSetup(self, setup):
        if (setup.get('name') == self.sSetupConfig):
            self.bSampleSetupSet(None)
        self.tXmlConfig.tree.find('Config').remove(setup)
        self.xmlSave()

    def newXmlSetup(self, setup, sConfig):
        cloneVersion = copy.deepcopy(setup)
        cloneVersion.set('name', sConfig)
        self.tXmlConfig.tree.find('Config').append(cloneVersion)
        self.xmlSave()
        self.bSampleSetupSet(sConfig)

    def tXmlChildNew(self, tParrent, sName):
        new = ET.SubElement(tParrent, sName)
        return new

    def xmlPrintSetups(self):
        for setup in self.tXmlConfig.tree.find('Config'):
            print(setup.get('name'))
            print("    Device Name: " + setup.find('DeviceName').text)
            print("    Acc: " + setup.find('Acc').text)
            iAcquisitionTime = AdcAcquisitionTime[int(
                setup.find('AcquisitionTime').text)]
            iOversampling = AdcOverSamplingRate[int(
                setup.find('OverSamples').text)]
            samplingRate = int(
                calcSamplingRate(int(setup.find('Prescaler').text),
                                 iAcquisitionTime, iOversampling) + 0.5)
            print(
                "    ADC Prescaler/AcquisitionTime/OversamplingRate(Samples/s): "
                + setup.find('Prescaler').text + "/" +
                setup.find('AcquisitionTime').text + "/" +
                setup.find('OverSamples').text + "(" + str(samplingRate) + ")")
            print("    ADC Reference Voltage: " + setup.find('AdcRef').text)
            print("    Log Name: " + setup.find('LogName').text)
            print("    RunTime/IntervalTime: " + setup.find('RunTime').text +
                  " " + setup.find('DisplayTime').text)
            print("    Display Time: " + setup.find('DisplayTime').text)

    def _vRunConsoleStartupShow(self):
        print("XML File: " + str(self.sXmlFileName))
        print("Product Configuration: " + str(self.sProduct) + " " +
              str(self.sConfig))
        print("Setup Configuration: " + str(self.sSetupConfig))
        print("AutoSave?: " + str(self.bSave))
        print("Table Calculation File: " + str(self.sSheetFile))
        print("Log Name: " + str(self.Can.Logger.filepath.name))
        print("Device Name (to be connected): " + str(self.sDevName))
        print("Bluetooth address(to be connected): " +
              str(self.iAddress))  # Todo machen
        print("AutoConnect?: " + str(self.bSthAutoConnect))
        print("Run Time: " + str(self.iRunTime) + "s")
        print("Interval Time: " + str(self.iIntervalTime) + "s")
        print("Display Time: " + str(self.iDisplayTime) + "s")
        print(
            "Adc Prescaler/AcquisitionTime/OversamplingRate/Reference(Samples/s): "
            + str(self.iPrescaler) + "/" +
            str(AdcAcquisitionTime.inverse[self.iAquistionTime]) + "/" +
            str(AdcOverSamplingRate.inverse[self.iOversampling]) + "/" +
            str(self.sAdcRef) + "(" + str(self.samplingRate) + ")")
        print("Acc Config(XYZ/DataSets): " + str(int(self.bAccX)) +
              str(int(self.bAccY)) + str(int(self.bAccZ)) + "/" +
              str(DataSets.inverse[self.tAccDataFormat]))
        print("Voltage Config(XYZ/DataSets): " + str(int(self.bVoltageX)) +
              str(int(self.bVoltageY)) + str(int(self.bVoltageZ)) + "/" +
              str(DataSets.inverse[self.tAccDataFormat]) + ("(X=Battery)"))

    def _vRunConsoleStartupLoggerPrint(self):
        self.Can.Logger.Info("XML File: " + str(self.sXmlFileName))
        self.Can.Logger.Info("Product Configuration: " + str(self.sProduct) +
                             " " + str(self.sConfig))
        self.Can.Logger.Info("Setup Configuration: " + str(self.sSetupConfig))
        self.Can.Logger.Info("AutoSave?: " + str(self.bSave))
        self.Can.Logger.Info("Table Calculation File: " + str(self.sSheetFile))
        self.Can.Logger.Info("Log Name: " + str(self.Can.Logger.filepath.name))
        self.Can.Logger.Info("Device Name (to be connected): " +
                             str(self.sDevName))
        self.Can.Logger.Info("Bluetooth address(to be connected): " +
                             str(self.iAddress))  # Todo machen
        self.Can.Logger.Info("AutoConnect?: " + str(self.bSthAutoConnect))
        self.Can.Logger.Info("Run Time: " + str(self.iRunTime) + "s")
        self.Can.Logger.Info("Interval Time: " + str(self.iIntervalTime) + "s")
        self.Can.Logger.Info("Display Time: " + str(self.iDisplayTime) + "ms")
        self.Can.Logger.Info(
            "Adc Prescaler/AcquisitionTime/OversamplingRate/Reference(Samples/s): "
            + str(self.iPrescaler) + "/" +
            str(AdcAcquisitionTime.inverse[self.iAquistionTime]) + "/" +
            str(AdcOverSamplingRate.inverse[self.iOversampling]) + "/" +
            str(self.sAdcRef) + "(" + str(self.samplingRate) + ")")
        self.Can.Logger.Info("Acc Config(XYZ/DataSets): " +
                             str(int(self.bAccX)) + str(int(self.bAccY)) +
                             str(int(self.bAccZ)) + "/" +
                             str(DataSets.inverse[self.tAccDataFormat]))
        self.Can.Logger.Info("Voltage Config(XYZ/DataSets): " +
                             str(int(self.bVoltageX)) +
                             str(int(self.bVoltageY)) +
                             str(int(self.bVoltageZ)) + "/" +
                             str(DataSets.inverse[self.tAccDataFormat]) +
                             ("(X=Battery)"))

    def _vRunConsoleStartup(self):
        self._vRunConsoleStartupLoggerPrint()
        if False != self.args_dict['show_config']:
            self._vRunConsoleStartupShow()
        if False != self.args_dict['show_products']:
            self.xmlPrintVersions()
        if False != self.args_dict['show_setups']:
            self.xmlPrintSetups()

    def clear(self):
        # for windows
        if os.name == 'nt':
            _ = os.system('cls')

        # for mac and linux(here, os.name is 'posix')
        else:
            _ = os.system('clear')

    def vRunConsoleAutoConnect(self):
        self.clear()
        if "0x0" != self.iAddress and 0 != self.iAddress and "0" != self.iAddress:
            self.Can.bBlueToothConnectPollingAddress(MyToolItNetworkNr["STU1"],
                                                     self.iAddress)
        else:
            self.Can.bBlueToothConnectPollingName(MyToolItNetworkNr["STU1"],
                                                  self.sDevName,
                                                  log=False)
        if False != self.Can.bConnected:
            self.vDataAquisition()

    def vRunConsole(self):
        self._vRunConsoleStartup()
        self.reset()
        if False != self.bSthAutoConnect:
            self.vRunConsoleAutoConnect()
        self.close()

if __name__ == "__main__":
    watch = myToolItWatch()
    watch.vParserInit()
    watch.vParserConsoleArgumentsPass()
    watch.vRunConsole()
