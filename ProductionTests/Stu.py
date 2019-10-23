import unittest
import os
import sys
sDirName = os.path.dirname('')
sys.path.append(sDirName)
file_path = '../'
sDirName = os.path.dirname(file_path)
sys.path.append(sDirName)
import CanFd
import time
import array
import openpyxl
import struct
import SthLimits
import MyToolItSth
from datetime import date
from shutil import copyfile
from openpyxl.styles import Font
from MyToolItNetworkNumbers import MyToolItNetworkNr 
from MyToolItStu import TestConfig, StuErrorWord

from StuLimits import PcbOnly, RssiStuMin
from MyToolItCommands import *
from openpyxl.descriptors.base import DateTime

sVersion = "v2.1.4"
sLogName = 'ProductionTestStu'
sLogLocation = '../../Logs/ProductionTestStu/'
sBuildLocation = "../../SimplicityStudio/v4_workspace/client_firmware/builds/"
sSilabsCommanderLocation = "../../SimplicityStudio/SimplicityCommander/"
sAdapterSerialNo = "440116697"
sBoardType = "BGM111A256V2"

def sSerialNumber(sExcelFileName):
    tWorkbook = openpyxl.load_workbook(sExcelFileName)
    tWorkSheet = tWorkbook.get_sheet_by_name("Product Data@0x4")
    value = tWorkSheet['E12'].value
    sSerial = str(value)
    return sSerial


bSkip = False


class TestStu(unittest.TestCase):

    def setUp(self):
        global bSkip
        self.sBuildLocation = sBuildLocation + sVersion
        self.sBootloader = sBuildLocation + "BootloaderOtaBgm111.s37"
        self.sAdapterSerialNo = sAdapterSerialNo
        self.sBoardType = sBoardType 
        self.sSilabsCommander = sSilabsCommanderLocation + "commander"
        self.iTestNumber = int(self._testMethodName[4:8])
        self.sLogLocation = sLogLocation
        self.fileName = self.sLogLocation + self._testMethodName + ".txt"
        self.fileNameError = self.sLogLocation + "Error_" + self._testMethodName + ".txt"
        self.sExcelEepromContentFileName = "Stu" + sVersion + ".xlsx"
        if False != bSkip and "test9999StoreTestResults" != self._testMethodName:
            self.skipTest("At least some previous test failed")
        else:
            self.Can = CanFd.CanFd(CanFd.PCAN_BAUD_1M, self.fileName, self.fileNameError, MyToolItNetworkNr["SPU1"], MyToolItNetworkNr["STU1"], 0, 0, 0, FreshLog=True)
            self.sSerialNumber = sSerialNumber(self.sExcelEepromContentFileName)
            self.sDateClock = sDateClock() 
            self.Can.Logger.Info("TestCase: " + str(self._testMethodName))
            if "test0000FirmwareFlash" != self._testMethodName:    
                self.Can.CanTimeStampStart(self._resetStu()["CanTime"])  # This will also reset to STH
                self.sStuAddr = sBlueToothMacAddr(self.Can.BlueToothAddress(MyToolItNetworkNr["STU1"]))
                self.sTestReport = sLogName + "_" + self.sSerialNumber + "_" + self.sStuAddr.replace(":", "#")
                sStoreFileName = "./ResultsStu/OK_" + self.sTestReport + "_nr0.xlsx"
                if "test0005Ack" == self._testMethodName:
                    batchFile = open("BatchNumberStu.txt", "w")
                    iBatchNr = int(self.sBatchNumber)
                    iBatchNr += 1
                    self.sBatchNumber = str(iBatchNr)
                    batchFile.write(self.sBatchNumber)
                    batchFile.close()
                self.tWorkbookOpenCreate()
                self.sExcelEepromContentReadBackFileName = sLogName + "_" + self.sSerialNumber + "_" + self.sStuAddr.replace(":", "#") + "_ReadBack.xlsx"
                self.Can.Logger.Info("STU BlueTooth Address: " + self.sStuAddr)
                self._statusWords()
                self._SthWDog()
            self.Can.Logger.Info("_______________________________________________________________________________________________________________")
            self.Can.Logger.Info("Start")

    def tearDown(self):
        global bSkip
        if "test9999StoreTestResults" != self._testMethodName and "test0000FirmwareFlash" != self._testMethodName:
            self.tWorkbook.save(self.sTestReport + ".xlsx")
        if False == self.Can.bError:  
            self.Can.Logger.Info("Fin")
            self.Can.Logger.Info("_______________________________________________________________________________________________________________")
        if "test0000FirmwareFlash" == self._testMethodName: 
            if False == self.Can.bError:  
                self.Can.CanTimeStampStart(self._resetStu()["CanTime"])  # This will also reset to STH
                self.sStuAddr = sBlueToothMacAddr(self.Can.BlueToothAddress(MyToolItNetworkNr["STU1"]))
                self.Can.Logger.Info("STU BlueTooth Address: " + self.sStuAddr)
                self._statusWords()
                self._SthWDog()
        if False == self.Can.bError:  
            self.Can.Logger.Info("Test Time End Time Stamp")
        self.Can.__exit__()
        if self._outcome.errors[1][1]:
            if False == bSkip:
                print("Error! Please put red point on it(" + self.sBatchNumber + ")")
            bSkip = True

    def run(self, result=None):
        global bSkip
        batchFile = open("BatchNumberStu.txt")
        sBatchData = batchFile.readlines()
        self.sBatchNumber = sBatchData[0]
        batchFile.close()
        """ Stop after first error """
        if not result.errors:             
            super(TestStu, self).run(result)
        else:
            if False == bSkip:
                print("Error! Please put red point on it(" + self.sBatchNumber + ")")
            bSkip = True 
            if "test9999StoreTestResults" == self._testMethodName:
                super(TestStu, self).run(result)

            
    def _resetStu(self, retries=5, log=True):
        self.Can.bConnected = False
        return self.Can.cmdReset(MyToolItNetworkNr["STU1"], retries=retries, log=log)
    
    def _statusWords(self):
        ErrorWord = StuErrorWord()
        psw0 = self.Can.statusWord0(MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STU Status Word: " + hex(psw0))
        ErrorWord.asword = self.Can.statusWord1(MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STU Error Word: " + hex(ErrorWord.asword))

    def _SthWDog(self):
        WdogCounter = iMessage2Value(self.Can.statisticalData(MyToolItNetworkNr["STU1"], MyToolItStatData["Wdog"])[:4])
        self.Can.Logger.Info("WatchDog Counter: " + str(WdogCounter))
        return WdogCounter 
           
    """
    Try to open Excel Sheet. If not able to open, then create a new one
    """

    def tWorkbookOpenCreate(self):
        try:
            self.tWorkbook = openpyxl.load_workbook(self.sTestReport + ".xlsx")
        except:
            self.tWorkbook = openpyxl.Workbook()      
            self.tWorkbook.remove_sheet(self.tWorkbook.get_sheet_by_name('Sheet'))
            self.tWorkSheet = self.tWorkbook.create_sheet(self.sStuAddr.replace(":", "#"))
            tFont = Font(bold=True, size=20)
            self.tWorkSheet['A1'] = 'Test Batch Number'
            self.tWorkSheet['A1'].font = tFont
            self.tWorkSheet['B1'] = 'Test Name'
            self.tWorkSheet['B1'].font = tFont
            self.tWorkSheet['C1'] = 'DateTime'
            self.tWorkSheet['C1'].font = tFont
            self.tWorkSheet['D1'] = 'Description'
            self.tWorkSheet['D1'].font = tFont
            for i in range(0, 16):
                self.tWorkSheet[chr(0x45 + i) + "1"].value = 'Test Case Report ' + str(i)
                self.tWorkSheet[chr(0x45 + i) + "1"].font = tFont
            self.tWorkbook.save(self.sTestReport + ".xlsx")
            self.tWorkbook = openpyxl.load_workbook(self.sTestReport + ".xlsx")
        self.tWorkSheet = self.tWorkbook.get_sheet_by_name(self.sStuAddr.replace(":", "#"))
        self.iTestRow = 1
        while(None != self.tWorkSheet['A' + str(self.iTestRow + 1)].value):
            self.iTestRow += 1
        self.tWorkSheetWrite("A", self.iTestRow)
        self.tWorkSheetWrite("B", self._testMethodName)
        self.tWorkSheetWrite("C", self.sDateClock)
    
    def tWorkSheetWrite(self, sCol, value):
        tFont = Font(bold=False, size=12)
        self.tWorkSheet[sCol + str(self.iTestRow + 1)].value = str(value)
        self.tWorkSheet[sCol + str(self.iTestRow + 1)].font = tFont

    def au8excelValueToByteArray(self, worksheet, iIndex):
        iLength = int(worksheet['C' + str(iIndex)].value)
        value = worksheet['E' + str(iIndex)].value
        byteArray = [0] * iLength
        if None != value:
            if "UTF8" == worksheet['G' + str(iIndex)].value:
                try:
                    value = str(value).encode('utf-8')
                except Exception as e: 
                    self.Can.Logger.Info(str(e))
                    value = [0] * iLength
                iCopyLength = len(value)
                if iLength < iCopyLength:
                    iCopyLength = iLength
                for i in range(0, iCopyLength):
                    byteArray[i] = value[i]
            elif "ASCII" == worksheet['G' + str(iIndex)].value:
                try:
                    value = str(value).encode('ascii')
                except Exception as e: 
                    self.Can.Logger.Info(str(e))
                    value = [0] * iLength
                iCopyLength = len(value)
                if iLength < iCopyLength:
                    iCopyLength = iLength
                for i in range(0, iCopyLength):
                    byteArray[i] = value[i]                
            elif "unsigned" == worksheet['G' + str(iIndex)].value:
                byteArray = au8Value2Array(int(value), iLength)   
            elif "float" == worksheet['G' + str(iIndex)].value:
                value = float(value)
                value = struct.pack('f', value)
                value = int.from_bytes(value, byteorder='little')
                byteArray = au8Value2Array(int(value), 4)   
            else:
                if "0" == value or 0 == value:
                    value = "[0x0]"
                value = value[1:-1].split(',')
                for i in range(0, len(value)):
                    byteArray[i] = int(value[i], 16)
        return byteArray
    
    def iExcelSheetPageLength(self, worksheet):
        totalLength = 0
        for i in range(2, 256 + 2):
            if None != worksheet['A' + str(i)].value:
                length = int(worksheet['C' + str(i)].value)
                totalLength += length
            else:
                break
        return totalLength 
    
    def sExcelSheetWrite(self, namePage, iReceiver):
        sError = None
        workbook = openpyxl.load_workbook(self.sExcelEepromContentFileName)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = int(name[1], base=16)
                name = name[0]
                if namePage == name:
                    worksheet = workbook.get_sheet_by_name(worksheetName)
                    # Prepare Write Data
                    au8WriteData = [0] * 256
                    iByteIndex = 0
                    for i in range(2, 256 + 2, 1):
                        if None != worksheet['A' + str(i)].value:
                            au8ElementData = self.au8excelValueToByteArray(worksheet, i)
                            for j in range(0, len(au8ElementData), 1):
                                au8WriteData[iByteIndex + j] = au8ElementData[j] 
                            iLength = int(worksheet['C' + str(i)].value)
                            iByteIndex += iLength
                        else:
                            break
                    iWriteLength = self.iExcelSheetPageLength(worksheet)  
                    if 0 != iWriteLength % 4:
                        iWriteLength += 4
                        iWriteLength -= (iWriteLength % 4)
                    au8WriteData = au8WriteData[0:iWriteLength] 
                    self.Can.Logger.Info("Write Content: " + payload2Hex(au8WriteData))
                    for offset in range(0, iWriteLength, 4):
                        au8WritePackage = au8WriteData[offset:offset + 4]
                        au8Payload = [address, 0xFF & offset, 4, 0]
                        au8Payload.extend(au8WritePackage)
                        self.Can.cmdSend(iReceiver, MyToolItBlock["Eeprom"], MyToolItEeprom["Write"], au8Payload, log=False)   
            try:
                workbook.close() 
            except Exception as e: 
                sError = "Could not close file: " + str(e)
                self.Can.Logger.Info(sError)
        return sError

    def vUnicodeIllegalRemove(self, value, character):
        while(True):
            try:
                value.remove(character)
            except:
                break  
        return value          
    
    def iExcelSheetPageValue(self, worksheet, aBytes):
        i = 2
        iTotalLength = 0
        while len(aBytes) > iTotalLength:
            if None != worksheet['A' + str(i)].value:
                iLength = worksheet['C' + str(i)].value
                value = aBytes[iTotalLength:iTotalLength + iLength]
                if "UTF8" == worksheet['G' + str(i)].value:
                    try:
                        value = self.vUnicodeIllegalRemove(value, 0)
                        value = self.vUnicodeIllegalRemove(value, 255)
                        value = array.array('b', value).tostring().decode('utf-8', 'replace')
                    except Exception as e:
                        self.Can.Logger.Info(str(e))
                        value = ""
                elif "ASCII" == worksheet['G' + str(i)].value:
                    value = sArray2String(value)
                elif "unsigned" == worksheet['G' + str(i)].value:
                    value = str(iMessage2Value(value))
                elif "float" == worksheet['G' + str(i)].value:
                    if None != value:
                        pass
                        # value = au8ChangeEndianOrder(value)
                    else: 
                        value = 0.0
                    self.Can.Logger.Info("Value from EEPROM: " + str(value))
                    value = bytearray(value)
                    value = struct.unpack('f', value)[0]
                    value = str(value)
                    self.Can.Logger.Info("Value as float: " + str(value))
                else:                    
                    value = payload2Hex(value)
                value = str(value)
                try:
                    worksheet['E' + str(i)] = str(value).encode('utf8')
                except:
                    pass
                iTotalLength += iLength
                i += 1
            else:
                break
        return iTotalLength  

    """
    Read EEPROM page to write values in Excel Sheet
    """    

    def sExcelSheetRead(self, namePage, iReceiver):
        sError = None
        workbook = openpyxl.load_workbook(self.sExcelEepromContentReadBackFileName)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = int(name[1], base=16)
                name = name[0]
                if namePage == name:
                    worksheet = workbook.get_sheet_by_name(worksheetName)
                    pageContent = []
                    readLength = self.iExcelSheetPageLength(worksheet)
                    readLengthAlligned = readLength
                    if 0 != readLengthAlligned % 4:
                        readLengthAlligned += 4
                        readLengthAlligned -= (readLengthAlligned % 4)
                    for offset in range(0, readLengthAlligned, 4):
                        payload = [address, 0xFF & offset, 4, 0, 0, 0, 0, 0]
                        index = self.Can.cmdSend(iReceiver, MyToolItBlock["Eeprom"], MyToolItEeprom["Read"], payload, log=False)   
                        readBackFrame = self.Can.getReadMessageData(index)[4:]
                        pageContent.extend(readBackFrame)
                    pageContent = pageContent[0:readLength]
                    self.Can.Logger.Info("Read Data: " + payload2Hex(pageContent))
                    self.iExcelSheetPageValue(worksheet, pageContent)
            try:
                workbook.save(self.sExcelEepromContentReadBackFileName) 
            except Exception as e: 
                sError = "Could not save file(Opened by another application?): " + str(e)
                self.Can.Logger.Info(sError)
        return sError
    
    def tCompareEerpomWriteRead(self):
        tWorkSheetNameError = None
        sCellNumberError = None
        tWorkbookReadBack = openpyxl.load_workbook(self.sExcelEepromContentReadBackFileName)
        tWorkbookWrite = openpyxl.load_workbook(self.sExcelEepromContentFileName)
        if tWorkbookReadBack and tWorkbookWrite:
            for worksheetName in tWorkbookWrite.sheetnames:
                tWorkSheedReadBack = tWorkbookReadBack.get_sheet_by_name(worksheetName)
                tWorkSheedWrite = tWorkbookWrite.get_sheet_by_name(worksheetName)
                for i in range(2, 2 + 256):
                    if None != tWorkSheedWrite['A' + str(i)].value:
                        if str(tWorkSheedWrite['E' + str(i)].value) != str(tWorkSheedReadBack['E' + str(i)].value):
                            tWorkSheetNameError = worksheetName
                            sCellNumberError = 'E' + str(i)
                            break
                    else:
                        if None != tWorkSheedReadBack['A' + str(i)].value:
                            tWorkSheetNameError = worksheetName
                            sCellNumberError = 'A' + str(i)
                        break
                            
        return [tWorkSheetNameError, sCellNumberError]
    
    """
    Change a specific content of an ExcelCell
    """
    
    def vChangeExcelCell(self, sWorkSheetName, sCellName, sContent):
        workSheetNames = []
        workbook = openpyxl.load_workbook(self.sExcelEepromContentFileName)
        if workbook:
            for worksheetName in workbook.sheetnames:
                sName = str(worksheetName).split('@')
                _sAddress = sName[1]
                sName = sName[0]
                workSheetNames.append(sName)
        worksheet = workbook.get_sheet_by_name(sWorkSheetName)    
        worksheet[sCellName] = sContent
        workbook.save(self.sExcelEepromContentFileName)   

    """
    https://www.silabs.com/community/wireless/zigbee-and-thread/knowledge-base.entry.html/2017/12/28/building_firmwareim-1OPr
    commander.exe convert ..\v4_workspace\client_firmware\builds\BootloaderOtaBgm111.s37 ..\v4_workspace\client_firmware\builds\v2.1.4\Client.s37 --patch 0x0fe04000:0x00 --patch 0x0fe041F8:0xFD -o manufacturing_image.hex -d BGM111A256V2 
    commander flash manufacturing_image.hex --address 0x0 --serialno 440116697 -d BGM111A256V2 
    """

    def test0000FirmwareFlash(self):      
        try:
            os.remove(self.sLogLocation + self._testMethodName + "ManufacturingCreateResport.txt")
        except:
            pass
        try:
            os.remove(self.sLogLocation + self._testMethodName + "ManufacturingFlashResport.txt")
        except:
            pass
        try:
            os.remove(self.sLogLocation + self._testMethodName + "ManufacturingDebugUnlock.txt")
        except:
            pass
        try:
            os.remove(self.sLogLocation + self._testMethodName + "DeviceInfo.txt")
        except:
            pass
        
        
        sSystemCall = self.sSilabsCommander + " device lock –-debug disable --serialno " + self.sAdapterSerialNo
        sSystemCall += " -d " + self.sBoardType
        sSystemCall += ">> " + self.sLogLocation 
        sSystemCall += self._testMethodName + "ManufacturingDebugUnlock.txt"
        if os.name == 'nt': 
            sSystemCall = sSystemCall.replace("/", "\\")
            os.system(sSystemCall)
        # for mac and linux(here, os.name is 'posix') 
        else: 
            os.system(sSystemCall)
        sSystemCall = self.sSilabsCommander + " device info "
        sSystemCall += "--serialno " + self.sAdapterSerialNo + " " 
        sSystemCall += "-d " + self.sBoardType + " "
        sSystemCall += ">> " + self.sLogLocation 
        sSystemCall += self._testMethodName + "DeviceInfo.txt"
        if os.name == 'nt': 
            sSystemCall = sSystemCall.replace("/", "\\")
            os.system(sSystemCall)
        # for mac and linux(here, os.name is 'posix') 
        else: 
            os.system(sSystemCall)
        tFile = open(self.sLogLocation + self._testMethodName + "DeviceInfo.txt", "r", encoding='utf-8')
        asData = tFile.readlines()
        tFile.close()            
        if "Unique ID" == asData[-2][:9]:          
            sSystemCall = self.sSilabsCommander + " convert "
            sSystemCall += self.sBootloader + " "
            sSystemCall += self.sBuildLocation + "/Client.s37 "
            sSystemCall += "--patch 0x0fe04000:0x00 --patch 0x0fe041F8:0xFD "
            sSystemCall += "-o " + self.sBuildLocation + "/manufacturing_imageStu" + sVersion + ".hex " 
            sSystemCall += "-d " + self.sBoardType + " "
            sSystemCall += ">> " + self.sLogLocation 
            sSystemCall += self._testMethodName + "ManufacturingCreateResport.txt"
            if os.name == 'nt': 
                sSystemCall = sSystemCall.replace("/", "\\")
                os.system(sSystemCall)
            # for mac and linux(here, os.name is 'posix') 
            else: 
                os.system(sSystemCall)
            tFile = open(self.sLogLocation + self._testMethodName + "ManufacturingCreateResport.txt", "r", encoding='utf-8')
            asData = tFile.readlines()
            tFile.close()
            self.assertEqual("Overwriting file:", asData[-2][:17])
            self.assertEqual("DONE\n", asData[-1])
            sSystemCall = self.sSilabsCommander + " flash "
            sSystemCall += self.sBuildLocation + "/manufacturing_imageStu" + sVersion + ".hex " 
            sSystemCall += "--address 0x0 "
            sSystemCall += "--serialno " + self.sAdapterSerialNo + " "
            sSystemCall += "-d " + self.sBoardType + " "
            sSystemCall += ">> " + self.sLogLocation 
            sSystemCall += self._testMethodName + "ManufacturingFlashResport.txt"
            if os.name == 'nt': 
                sSystemCall = sSystemCall.replace("/", "\\")
                os.system(sSystemCall)
            # for mac and linux(here, os.name is 'posix') 
            else: 
                os.system(sSystemCall)    
            tFile = open(self.sLogLocation + self._testMethodName + "ManufacturingFlashResport.txt", "r", encoding='utf-8')
            asData = tFile.readlines()
            tFile.close()
            self.assertEqual("range 0x0FE04000 - 0x0FE047FF (2 KB)\n", asData[-2][10:])   
            self.assertEqual("DONE\n", asData[-1])    

                       
    """
    Test Acknowledgement from STH. Write message and check identifier to be ack (No bError)
    """

    def test0005Ack(self):
        self.tWorkSheetWrite("D", "This test case checks the ability to communicate with the STU")
        cmd = self.Can.CanCmd(MyToolItBlock["System"], MyToolItSystem["ActiveState"], 1, 0)
        expectedData = ActiveState()
        expectedData.asbyte = 0
        expectedData.b.u2NodeState = Node["Application"]
        expectedData.b.u3NetworkState = NetworkState["Operating"]
        msg = self.Can.CanMessage20(cmd, MyToolItNetworkNr["SPU1"], MyToolItNetworkNr["STU1"], [expectedData.asbyte])
        self.Can.Logger.Info("Write Message")
        self.Can.WriteFrame(msg)
        self.Can.Logger.Info("Wait 200ms")
        time.sleep(0.2)
        cmd = self.Can.CanCmd(MyToolItBlock["System"], MyToolItSystem["ActiveState"], 0, 0)
        msgAckExpected = self.Can.CanMessage20(cmd, MyToolItNetworkNr["STU1"], MyToolItNetworkNr["SPU1"], [0])
        self.Can.Logger.Info("Send ID: " + hex(msg.ID) + "; Expected ID: " + hex(msgAckExpected.ID) + "; Received ID: " + hex(self.Can.getReadMessage(-1).ID))
        self.Can.Logger.Info("Send Data: " + hex(0) + "; Expected Data: " + hex(expectedData.asbyte) + "; Received Data: " + hex(self.Can.getReadMessage(-1).DATA[0]))
        self.tWorkSheetWrite("E", "Test failed")
        self.assertEqual(hex(msgAckExpected.ID), hex(self.Can.getReadMessage(-1).ID))
        self.assertEqual(expectedData.asbyte, self.Can.getReadMessage(-1).DATA[0])
        self.tWorkSheetWrite("E", "Test OK")
        
    """
    Checks that correct Firmware Version has been installed
    """

    def test0040Version(self):
        self.tWorkSheetWrite("D", "Check that the correct firmware version has been installed")
        iIndex = self.Can.cmdSend(MyToolItNetworkNr["STU1"], MyToolItBlock["ProductData"], MyToolItProductData["FirmwareVersion"], [])
        au8Version = self.Can.getReadMessageData(iIndex)[-3:]
        sVersionRead = "v" + str(au8Version[0]) + "." + str(au8Version[1]) + "." + str(au8Version[2])
        if sVersionRead == sVersion:
            self.tWorkSheetWrite("E", "OK")
        else:
            self.tWorkSheetWrite("E", "NOK")
        self.assertEqual(sVersionRead, sVersion)
    
    """
    Test Reset
    """   

    def test0099Reset(self):
        self.tWorkSheetWrite("D", "Tests Reset Command")
        self._resetStu()
        cmd = self.Can.CanCmd(MyToolItBlock["System"], MyToolItSystem["ActiveState"], 1, 0)
        expectedData = ActiveState()
        expectedData.asbyte = 0
        expectedData.b.u2NodeState = Node["Application"]
        expectedData.b.u3NetworkState = NetworkState["Operating"]
        msg = self.Can.CanMessage20(cmd, MyToolItNetworkNr["SPU1"], MyToolItNetworkNr["STU1"], [expectedData.asbyte])
        self.Can.Logger.Info("Write Message")
        self.Can.WriteFrame(msg)
        self.Can.Logger.Info("Wait 200ms")
        time.sleep(0.2)
        cmd = self.Can.CanCmd(MyToolItBlock["System"], MyToolItSystem["ActiveState"], 0, 0)
        msgAckExpected = self.Can.CanMessage20(cmd, MyToolItNetworkNr["STU1"], MyToolItNetworkNr["SPU1"], [0])
        self.Can.Logger.Info("Send ID: " + hex(msg.ID) + "; Expected ID: " + hex(msgAckExpected.ID) + "; Received ID: " + hex(self.Can.getReadMessage(-1).ID))
        self.Can.Logger.Info("Send Data: " + hex(0) + "; Expected Data: " + hex(expectedData.asbyte) + "; Received Data: " + hex(self.Can.getReadMessage(-1).DATA[0]))
        self.tWorkSheetWrite("E", "Test failed")
        self.assertEqual(hex(msgAckExpected.ID), hex(self.Can.getReadMessage(-1).ID))
        self.assertEqual(expectedData.asbyte, self.Can.getReadMessage(-1).DATA[0])
        self.tWorkSheetWrite("E", "Test OK")
        
    """
    Test that RSSI is good enough
    """

    def test0100Rssi(self):
        self.tWorkSheetWrite("D", "Tests RSSI")
        self.Can.bBlueToothConnectPollingName(MyToolItNetworkNr["STU1"], MyToolItSth.TestConfig["DevName"], log=False)
        time.sleep(1)
        self.sStuAddr = sBlueToothMacAddr(self.Can.BlueToothAddress(MyToolItNetworkNr["STU1"]))
        self.sSthAddr = sBlueToothMacAddr(self.Can.BlueToothAddress(MyToolItNetworkNr["STH1"]))
        self.Can.Logger.Info("STH BlueTooth Address: " + self.sSthAddr)
        iRssiSth = int(self.Can.BlueToothRssi(MyToolItNetworkNr["STH1"]))
        iRssiStu = int(self.Can.BlueToothRssi(MyToolItNetworkNr["STU1"]))
        self.tWorkSheetWrite("E", "RSSI @ STH: " + str(iRssiSth) + "dBm")
        self.tWorkSheetWrite("F", "RSSI @ STU: " + str(iRssiStu) + "dBm")
        self.Can.bBlueToothDisconnect(MyToolItNetworkNr["STU1"])
        self.assertGreater(iRssiSth, SthLimits.RssiSthMin)
        self.assertLess(iRssiSth, -20)
        self.assertGreater(iRssiStu, RssiStuMin)
        self.assertLess(iRssiStu, -20)
         
    """
    Write EEPROM
    """

    def test0399Eerpom(self):
        self._resetStu()
        self.tWorkSheetWrite("D", "Write EEPROM with data and check that by read")
        self.vChangeExcelCell("Statistics@0x5", "E8", self.sBatchNumber)
        # Write
        workSheetNames = []
        workbook = openpyxl.load_workbook(self.sExcelEepromContentFileName)
        if workbook:
            for worksheetName in workbook.sheetnames:
                sName = str(worksheetName).split('@')
                _sAddress = sName[1]
                sName = sName[0]
                workSheetNames.append(sName)  
        sDate = date.today()
        sDate = str(sDate).replace('-', '')
        self.vChangeExcelCell("Statistics@0x5", "E7", sDate)       
        for pageName in workSheetNames:
            sError = self.sExcelSheetWrite(pageName, MyToolItNetworkNr["STU1"])  
            if None != sError:
                break      
        # Read Back
        sError = None
        copyfile(self.sExcelEepromContentFileName, self.sExcelEepromContentReadBackFileName)
        for pageName in workSheetNames:
            sError = self.sExcelSheetRead(pageName, MyToolItNetworkNr["STU1"])
            if None != sError:
                break
        [tWorkSheetNameError, sCellNumberError] = self.tCompareEerpomWriteRead()
        self.tWorkSheetWrite("E", "Error Worksheet: " + str(tWorkSheetNameError))
        self.tWorkSheetWrite("F", "Error Cell: " + str(sCellNumberError))
        self.assertEqual(tWorkSheetNameError, None)
                
    """
    Move Test Report to Results
    """

    def test9999StoreTestResults(self):
        global bSkip
        self.tWorkSheetWrite("D", "Store Results")
        if False != os.path.isfile(self.sExcelEepromContentReadBackFileName):
            tWorkbookContent = openpyxl.load_workbook(self.sExcelEepromContentReadBackFileName)
            for worksheetName in tWorkbookContent.sheetnames:
                tWorkSheet = self.tWorkbook.create_sheet(worksheetName)
                tWorkSheetContent = tWorkbookContent.get_sheet_by_name(worksheetName)
                for row in tWorkSheetContent:
                    for cell in row:
                        tWorkSheet[cell.coordinate].value = cell.value
            tWorkbookContent.close()
            os.remove(self.sExcelEepromContentReadBackFileName)
        if False != bSkip:
            self.tWorkSheetWrite("E", "NOK")   
            self.tWorkbook.save(self.sTestReport + ".xlsx")  
            for i in range(0, 100):
                sStoreFileName = self.sLogLocation + "/ResultsStu/NOK_" + self.sTestReport + "_nr" + str(i) + ".xlsx"
                if False == os.path.isfile(sStoreFileName):
                    os.rename(self.sTestReport + ".xlsx", sStoreFileName)    
                    break                
        else:
            self.tWorkSheetWrite("E", "OK")
            self.tWorkbook.save(self.sTestReport + ".xlsx")
            for i in range(0, 100):
                sStoreFileName = self.sLogLocation + "/ResultsStu/OK_" + self.sTestReport + "_nr" + str(i) + ".xlsx"
                if False == os.path.isfile(sStoreFileName):
                    os.rename(self.sTestReport + ".xlsx", sStoreFileName) 
                    break

                
if __name__ == "__main__":
    sLogLocation = sys.argv[1]
    sLogFile = sys.argv[2]
    if '/' != sLogLocation[-1]:
        sLogLocation += '/'
    sLogFileLocation = sLogLocation + sLogFile
    sLogLocation = sLogLocation
    sVersion = sys.argv[3]
    sDirName = os.path.dirname(sLogFileLocation)
    sys.path.append(sDirName)

    if not os.path.exists(sDirName):
        os.makedirs(sDirName)
    with open(sLogFileLocation, "w") as f:    
        runner = unittest.TextTestRunner(f)
        unittest.main(argv=['first-arg-is-ignored'], testRunner=runner)  
