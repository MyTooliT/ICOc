import unittest
import os
import sys
sDirName = os.path.dirname('')
sys.path.append(sDirName)
file_path = '../'
sDirName = os.path.dirname(file_path)
sys.path.append(sDirName)
import CanFd
import time
import array
import openpyxl
import math
import struct
from datetime import date
from shutil import copyfile
from openpyxl.styles import Font
from MyToolItNetworkNumbers import MyToolItNetworkNr
from MyToolItSth import TestConfig, SthErrorWord, fVoltageBattery
from MyToolItStu import StuErrorWord
from SthLimits import SthLimits
from StuLimits import StuLimits
from MyToolItCommands import *
from openpyxl.descriptors.base import DateTime
from os.path import abspath, isfile, join
from re import escape, search
from subprocess import run

sVersion = "v2.1.9"
sLogName = 'ProductionTestSth'
sLogLocation = '../../Logs/ProductionTestSth/'
sOtaComPort = 'COM6'
sBuildLocation = "../../STH/builds/"
sSilabsCommanderLocation = "C:/SiliconLabs/SimplicityStudio/v4/developer/adapter_packs/commander/"
sAdapterSerialNo = "440120910"
sBoardType = "BGM113A256V2"
iSensorAxis = 1
uAdc2Acc = 200
iRssiMin = -75
bStuPcbOnly = True
"""
Get serial number that is stored in the excel file
"""


def sSerialNumber(sExcelFileName):
    tWorkbook = openpyxl.load_workbook(sExcelFileName)
    tWorkSheet = tWorkbook["Product Data@0x4"]
    value = tWorkSheet['E12'].value
    sSerial = str(value)
    return sSerial


bSkip = False
sHolderNameInput = None
"""
This class supports a production test of the Sensory Tool Holder (STH)
"""


class TestSth(unittest.TestCase):
    def setUp(self):
        global bSkip
        global sHolderNameInput
        self.tSthLimits = SthLimits(iSensorAxis, uAdc2Acc, 20, 35)
        self.tStuLimits = StuLimits(bStuPcbOnly)
        self.sBuildLocation = sBuildLocation + sVersion
        self.sBootloader = sBuildLocation + "BootloaderOtaBgm113.s37"
        self.sAdapterSerialNo = sAdapterSerialNo
        self.sBoardType = sBoardType
        self.sSilabsCommander = sSilabsCommanderLocation + "commander"
        self.sLogLocation = sLogLocation
        self.bError = False
        self.sBuildLocation = sBuildLocation + sVersion
        self.iTestNumber = int(self._testMethodName[4:8])
        self.fileName = self.sLogLocation + self._testMethodName + ".txt"
        self.fileNameError = self.sLogLocation + "Error_" + self._testMethodName + ".txt"
        self.sExcelEepromContentFileName = "Sth" + sVersion + ".xlsx"

        if False != bSkip and "test9999StoreTestResults" != self._testMethodName:
            self.skipTest("At least some previous test failed")
        else:
            self.sDateClock = sDateClock()
            self.Can = CanFd.CanFd(
                CanFd.PCAN_BAUD_1M,
                self.fileName,
                self.fileNameError,
                MyToolItNetworkNr["SPU1"],
                MyToolItNetworkNr["STH1"],
                self.tSthLimits.uSamplingRatePrescalerReset,
                self.tSthLimits.uSamplingRateAcqTimeReset,
                self.tSthLimits.uSamplingRateOverSamplesReset,
                FreshLog=True)
            self.Can.Logger.Info("TestCase: " + str(self._testMethodName))
            self.sSerialNumber = sSerialNumber(
                self.sExcelEepromContentFileName)
            self.Can.CanTimeStampStart(
                self._resetStu()["CanTime"])  # This will also reset the STH

            if "test0000FirmwareFlash" != self._testMethodName:
                self.Can.Logger.Info("Connect to STH")
                self.sHolderName = TestConfig["DevName"]
                if None != self.sHolderName:
                    if "test0001OverTheAirUpdate" == self._testMethodName:
                        for _i in range(0, 2):
                            atDevList = self.Can.tDeviceList(
                                MyToolItNetworkNr["STU1"])
                            for tDev in atDevList:
                                if tDev["Name"] == sHolderNameInput:
                                    self.sHolderName = sHolderNameInput
                                    break
                            if self.sHolderName == sHolderNameInput:
                                break
                            time.sleep(2)
                    else:
                        self.sHolderName = sHolderNameInput
                self.Can.bBlueToothConnectPollingName(
                    MyToolItNetworkNr["STU1"], self.sHolderName, log=False)
                time.sleep(2)
                self.sSthAddr = sBlueToothMacAddr(
                    self.Can.BlueToothAddress(MyToolItNetworkNr["STH1"]))
                self.sTestReport = sHolderNameInput + "_" + sLogName
                sStoreFileName = "./ResultsSth/OK_" + self.sTestReport + "_nr0.xlsx"
                if False == os.path.isfile(
                        sStoreFileName
                ) and "test0001OverTheAirUpdate" == self._testMethodName:
                    batchFile = open("BatchNumberSth.txt", "w")
                    iBatchNr = int(self.sBatchNumber)
                    iBatchNr += 1
                    self.sBatchNumber = str(iBatchNr)
                    batchFile.write(self.sBatchNumber)
                    batchFile.close()
                self.sExcelEepromContentReadBackFileName = sLogName + "_" + self.sSerialNumber + "_" + self.sSthAddr.replace(
                    ":", "#") + "_ReadBack.xlsx"
                self.tWorkbookOpenCreate()
                self._statusWords()
                self._iSthAdcTemp()
                self._SthWDog()
                self.Can.Logger.Info("STH BlueTooth Address: " + self.sSthAddr)
            else:
                sHolderNameInput = "Blubb"
                if 8 < len(sHolderNameInput):
                    sHolderNameInput = sHolderName[:8]
            self.sStuAddr = sBlueToothMacAddr(
                self.Can.BlueToothAddress(MyToolItNetworkNr["STU1"]))
            self.Can.Logger.Info("STU BlueTooth Address: " + self.sStuAddr)
            self.Can.Logger.Info(
                "_______________________________________________________________________________________________________________"
            )
            self.Can.Logger.Info("Start")

    def tearDown(self):
        global bSkip
        self.Can.Logger.Info("Fin")
        self.Can.Logger.Info(
            "_______________________________________________________________________________________________________________"
        )
        if "test9999StoreTestResults" != self._testMethodName and "test0000FirmwareFlash" != self._testMethodName:
            self.tWorkbook.save(self.sTestReport + ".xlsx")
        if "test0000FirmwareFlash" != self._testMethodName and "test9999StoreTestResults" != self._testMethodName:
            self.Can.streamingStop(MyToolItNetworkNr["STH1"],
                                   MyToolItStreaming["Acceleration"])
            self.Can.streamingStop(MyToolItNetworkNr["STH1"],
                                   MyToolItStreaming["Voltage"])
            self._statusWords()
            self._iSthAdcTemp()
            self.Can.Logger.Info("Test Time End Time Stamp")
        self.Can.bBlueToothDisconnect(MyToolItNetworkNr["STU1"])
        self.Can.__exit__()
        if self._outcome.errors[1][1]:
            if False == bSkip:
                print("Error! Please put red point on it(" +
                      self.sBatchNumber + ")")
            bSkip = True
        if False != self.bError:
            if False == bSkip:
                print("Error! Please put red point on it(" +
                      self.sBatchNumber + ")")
            bSkip = True

    def run(self, result=None):
        global bSkip
        batchFile = open("BatchNumberSth.txt")
        sBatchData = batchFile.readlines()
        self.sBatchNumber = sBatchData[0]
        batchFile.close()
        """ Stop after first error """
        if not result.errors:
            super(TestSth, self).run(result)
        else:
            if False == bSkip:
                print("Error! Please put red point on it(" +
                      self.sBatchNumber + ")")
            bSkip = True
            if "test9999StoreTestResults" == self._testMethodName:
                super(TestSth, self).run(result)

    """
    Reset STU
    """

    def _resetStu(self, retries=5, log=True):
        self.Can.bConnected = False
        return self.Can.cmdReset(MyToolItNetworkNr["STU1"],
                                 retries=retries,
                                 log=log)

    """
    Get internal BGM113 Chip Temperature in °C of STH
    """

    def _iSthAdcTemp(self):
        ret = self.Can.calibMeasurement(MyToolItNetworkNr["STH1"],
                                        CalibMeassurementActionNr["Measure"],
                                        CalibMeassurementTypeNr["Temp"],
                                        1,
                                        AdcReference["1V25"],
                                        log=False)
        result = float(iMessage2Value(ret[4:]))
        result /= 1000
        self.Can.Logger.Info("Temperature(Chip): " + str(result) + "°C")
        self.Can.calibMeasurement(MyToolItNetworkNr["STH1"],
                                  CalibMeassurementActionNr["None"],
                                  CalibMeassurementTypeNr["Temp"],
                                  1,
                                  AdcReference["VDD"],
                                  log=False,
                                  bReset=True)
        return result

    """
    Get all status words of STH and STU
    """

    def _statusWords(self):
        ErrorWordSth = SthErrorWord()
        ErrorWordStu = StuErrorWord()
        psw0 = self.Can.statusWord0(MyToolItNetworkNr["STH1"])
        self.Can.Logger.Info("STH Status Word: " + hex(psw0))
        psw0 = self.Can.statusWord0(MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STU Status Word: " + hex(psw0))
        ErrorWordSth.asword = self.Can.statusWord1(MyToolItNetworkNr["STH1"])
        if True == ErrorWordSth.b.bAdcOverRun:
            self.bError = True
        self.Can.Logger.Info("STH Error Word: " + hex(ErrorWordSth.asword))
        ErrorWordStu.asword = self.Can.statusWord1(MyToolItNetworkNr["STU1"])
        self.Can.Logger.Info("STU Error Word: " + hex(ErrorWordStu.asword))

    """
    Retrieve STH Watchdog counter
    """

    def _SthWDog(self):
        WdogCounter = iMessage2Value(
            self.Can.statisticalData(MyToolItNetworkNr["STH1"],
                                     MyToolItStatData["Wdog"])[:4])
        self.Can.Logger.Info("WatchDog Counter: " + str(WdogCounter))
        return WdogCounter

    """
    Try to open Excel Sheet. If not able to open, then create a new one
    """

    def tWorkbookOpenCreate(self):
        try:
            self.tWorkbook = openpyxl.load_workbook(self.sTestReport + ".xlsx")
        except:
            self.tWorkbook = openpyxl.Workbook()
            self.tWorkbook.remove_sheet(
                self.tWorkbook.get_sheet_by_name('Sheet'))
            self.tWorkSheet = self.tWorkbook.create_sheet(
                self.sSthAddr.replace(":", "#"))
            tFont = Font(bold=True, size=20)
            self.tWorkSheet['A1'] = 'Test Batch Number'
            self.tWorkSheet['A1'].font = tFont
            self.tWorkSheet['B1'] = 'Test Name'
            self.tWorkSheet['B1'].font = tFont
            self.tWorkSheet['C1'] = 'DateTime'
            self.tWorkSheet['C1'].font = tFont
            self.tWorkSheet['D1'] = 'Description'
            self.tWorkSheet['D1'].font = tFont
            for i in range(0, 16):
                self.tWorkSheet[chr(0x45 + i) +
                                "1"].value = 'Test Case Report ' + str(i)
                self.tWorkSheet[chr(0x45 + i) + "1"].font = tFont
            self.tWorkbook.save(self.sTestReport + ".xlsx")
            self.tWorkbook = openpyxl.load_workbook(self.sTestReport + ".xlsx")
        self.tWorkSheet = self.tWorkbook.get_sheet_by_name(
            self.sSthAddr.replace(":", "#"))
        self.iTestRow = 1
        while (None != self.tWorkSheet['A' + str(self.iTestRow + 1)].value):
            self.iTestRow += 1
        self.tWorkSheetWrite("A", self.iTestRow)
        self.tWorkSheetWrite("B", self._testMethodName)
        self.tWorkSheetWrite("C", self.sDateClock)

    """
    Write Column in Row by value (Note that sColl is e.g. "D")
    """

    def tWorkSheetWrite(self, sCol, value):
        tFont = Font(bold=False, size=12)
        self.tWorkSheet[sCol + str(self.iTestRow + 1)].value = str(value)
        self.tWorkSheet[sCol + str(self.iTestRow + 1)].font = tFont

    """
    Transform excel sheet data to byte arrays such that the firmware can write
    it. Please take always care of byte endian style i.e. MSB vs LSB.
    """

    def au8excelValueToByteArray(self, worksheet, iIndex):
        iLength = int(worksheet['C' + str(iIndex)].value)
        value = worksheet['E' + str(iIndex)].value
        byteArray = [0] * iLength
        if None != value:
            if "UTF8" == worksheet['G' + str(iIndex)].value:
                try:
                    value = str(value).encode('utf-8')
                except Exception as e:
                    self.Can.Logger.Info(str(e))
                    value = [0] * iLength
                iCopyLength = len(value)
                if iLength < iCopyLength:
                    iCopyLength = iLength
                for i in range(0, iCopyLength):
                    byteArray[i] = value[i]
            elif "ASCII" == worksheet['G' + str(iIndex)].value:
                try:
                    value = str(value).encode('ascii')
                except Exception as e:
                    self.Can.Logger.Info(str(e))
                    value = [0] * iLength
                iCopyLength = len(value)
                if iLength < iCopyLength:
                    iCopyLength = iLength
                for i in range(0, iCopyLength):
                    byteArray[i] = value[i]
            elif "unsigned" == worksheet['G' + str(iIndex)].value:
                byteArray = au8Value2Array(int(value), iLength)
            elif "float" == worksheet['G' + str(iIndex)].value:
                value = float(value)
                value = struct.pack('f', value)
                value = int.from_bytes(value, byteorder='little')
                byteArray = au8Value2Array(int(value), 4)
            else:
                if "0" == value or 0 == value:
                    value = "[0x0]"
                value = value[1:-1].split(',')
                for i in range(0, len(value)):
                    byteArray[i] = int(value[i], 16)
        return byteArray

    """
    How many bytes are represented in Excel Sheet
    """

    def iExcelSheetPageLength(self, worksheet):
        totalLength = 0
        for i in range(2, 256 + 2):
            if None != worksheet['A' + str(i)].value:
                length = int(worksheet['C' + str(i)].value)
                totalLength += length
            else:
                break
        return totalLength

    """
    Read Excel Sheet and write it to STH
    """

    def sExcelSheetWrite(self, namePage, iReceiver):
        sError = None
        workbook = openpyxl.load_workbook(self.sExcelEepromContentFileName)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = int(name[1], base=16)
                name = name[0]
                if namePage == name:
                    worksheet = workbook.get_sheet_by_name(worksheetName)
                    # Prepare Write Data
                    au8WriteData = [0] * 256
                    iByteIndex = 0
                    for i in range(2, 256 + 2, 1):
                        if None != worksheet['A' + str(i)].value:
                            au8ElementData = self.au8excelValueToByteArray(
                                worksheet, i)
                            for j in range(0, len(au8ElementData), 1):
                                au8WriteData[iByteIndex +
                                             j] = au8ElementData[j]
                            iLength = int(worksheet['C' + str(i)].value)
                            iByteIndex += iLength
                        else:
                            break
                    iWriteLength = self.iExcelSheetPageLength(worksheet)
                    if 0 != iWriteLength % 4:
                        iWriteLength += 4
                        iWriteLength -= (iWriteLength % 4)
                    au8WriteData = au8WriteData[0:iWriteLength]
                    self.Can.Logger.Info("Write Content: " +
                                         payload2Hex(au8WriteData))
                    for offset in range(0, iWriteLength, 4):
                        au8WritePackage = au8WriteData[offset:offset + 4]
                        au8Payload = [address, 0xFF & offset, 4, 0]
                        au8Payload.extend(au8WritePackage)
                        self.Can.cmdSend(iReceiver,
                                         MyToolItBlock["Eeprom"],
                                         MyToolItEeprom["Write"],
                                         au8Payload,
                                         log=False)
            try:
                workbook.close()
            except Exception as e:
                sError = "Could not close file: " + str(e)
                self.Can.Logger.Info(sError)
        return sError

    """
    Try to get of an illegal UTF8 character
    """

    def vUnicodeIllegalRemove(self, value, character):
        while (True):
            try:
                value.remove(character)
            except:
                break
        return value

    """
    Write Byte Array into Excel Sheet with care of the format
    """

    def iExcelSheetPageValue(self, worksheet, aBytes):
        i = 2
        iTotalLength = 0
        while len(aBytes) > iTotalLength:
            if None != worksheet['A' + str(i)].value:
                iLength = worksheet['C' + str(i)].value
                value = aBytes[iTotalLength:iTotalLength + iLength]
                if "UTF8" == worksheet['G' + str(i)].value:
                    try:
                        value = self.vUnicodeIllegalRemove(value, 0)
                        value = self.vUnicodeIllegalRemove(value, 255)
                        value = array.array('b', value).tostring().decode(
                            'utf-8', 'replace')
                    except Exception as e:
                        self.Can.Logger.Info(str(e))
                        value = ""
                elif "ASCII" == worksheet['G' + str(i)].value:
                    value = sArray2String(value)
                elif "unsigned" == worksheet['G' + str(i)].value:
                    value = str(iMessage2Value(value))
                elif "float" == worksheet['G' + str(i)].value:
                    if None != value:
                        pass
                        # value = au8ChangeEndianOrder(value)
                    else:
                        value = 0.0
                    self.Can.Logger.Info("Value from EEPROM: " + str(value))
                    value = bytearray(value)
                    value = struct.unpack('f', value)[0]
                    value = str(value)
                    self.Can.Logger.Info("Value as float: " + str(value))
                else:
                    value = payload2Hex(value)
                value = str(value)
                try:
                    worksheet['E' + str(i)] = str(value).encode('utf8')
                except:
                    pass
                iTotalLength += iLength
                i += 1
            else:
                break
        return iTotalLength

    """
    Read EEPROM page to write values in Excel Sheet
    """

    def sExcelSheetRead(self, namePage, iReceiver):
        sError = None
        workbook = openpyxl.load_workbook(
            self.sExcelEepromContentReadBackFileName)
        if workbook:
            for worksheetName in workbook.sheetnames:
                name = str(worksheetName).split('@')
                address = int(name[1], base=16)
                name = name[0]
                if namePage == name:
                    worksheet = workbook.get_sheet_by_name(worksheetName)
                    pageContent = []
                    readLength = self.iExcelSheetPageLength(worksheet)
                    readLengthAlligned = readLength
                    if 0 != readLengthAlligned % 4:
                        readLengthAlligned += 4
                        readLengthAlligned -= (readLengthAlligned % 4)
                    for offset in range(0, readLengthAlligned, 4):
                        payload = [address, 0xFF & offset, 4, 0, 0, 0, 0, 0]
                        index = self.Can.cmdSend(iReceiver,
                                                 MyToolItBlock["Eeprom"],
                                                 MyToolItEeprom["Read"],
                                                 payload,
                                                 log=False)
                        readBackFrame = self.Can.getReadMessageData(index)[4:]
                        pageContent.extend(readBackFrame)
                    pageContent = pageContent[0:readLength]
                    self.Can.Logger.Info("Read Data: " +
                                         payload2Hex(pageContent))
                    self.iExcelSheetPageValue(worksheet, pageContent)
            try:
                workbook.save(self.sExcelEepromContentReadBackFileName)
            except Exception as e:
                sError = "Could not save file(Opened by another application?): " + str(
                    e)
                self.Can.Logger.Info(sError)
        return sError

    def test0000FirmwareFlash(self):
        """Upload bootloader into STH"""

        identification_arguments = (f"--serialno {self.sAdapterSerialNo} " +
                                    f"-d {self.sBoardType}")

        # Unlock debug access
        unlock_command = (
            f"{self.sSilabsCommander} device unlock {identification_arguments}"
        )
        if os.name == 'nt':
            unlock_command = unlock_command.replace('/', '\\')
        status = run(unlock_command, capture_output=True, text=True)
        self.assertEqual(
            status.returncode, 0,
            f"Unlock command returned non-zero exit code {status.returncode}")
        self.assertRegex(status.stdout, "Chip successfully unlocked",
                         "Unable to unlock debug access of chip")

        # Retrieve device id
        info_command = (
            f"{self.sSilabsCommander} device info {identification_arguments}")
        if os.name == 'nt':
            info_command = info_command.replace('/', '\\')
        status = run(info_command, capture_output=True, text=True)
        self.assertEqual(
            status.returncode, 0,
            "Device information command returned non-zero exit code " +
            f"{status.returncode}")
        id_match = search("Unique ID\s*:\s*(?P<id>[0-9A-Fa-f]+)",
                          status.stdout)
        self.assertIsNotNone(id_match, "Unable to determine unique ID of chip")
        unique_id = id_match['id']

        # Upload bootloader data
        bootloader_filepath = abspath(
            join(sBuildLocation, sVersion,
                 f"manufacturingImageSth{sVersion}.hex"))
        self.assertTrue(
            isfile(bootloader_filepath),
            f"Bootloader file {bootloader_filepath} does not exist")

        flash_command = (
            f"{self.sSilabsCommander} flash {bootloader_filepath} " +
            f"--address 0x0 {identification_arguments}")
        if os.name == 'nt':
            flash_command = flash_command.replace('/', '\\')
        status = run(flash_command, capture_output=True, text=True)
        self.assertEqual(
            status.returncode, 0,
            "Flash program command returned non-zero exit code " +
            f"{status.returncode}")
        expected_output = "range 0x0FE04000 - 0x0FE047FF (2 KB)"
        self.assertRegex(
            status.stdout, escape(expected_output),
            f"Flash output did not contain expected output “{expected_output}”"
        )
        expected_output = "DONE"
        self.assertRegex(
            status.stdout, expected_output,
            f"Flash output did not contain expected output “{expected_output}”"
        )


if __name__ == "__main__":
    sLogLocation = sys.argv[1]
    sLogFile = sys.argv[2]
    if '/' != sLogLocation[-1]:
        sLogLocation += '/'
    sLogFileLocation = sLogLocation + sLogFile
    sLogLocation = sLogLocation
    sVersion = sys.argv[3]
    sDirName = os.path.dirname(sLogFileLocation)
    sys.path.append(sDirName)

    if not os.path.exists(sDirName):
        os.makedirs(sDirName)
    with open(sLogFileLocation, "w") as f:
        unittest.main(argv=['first-arg-is-ignored'])
